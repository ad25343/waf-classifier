"""
WAF Category Classifier - Backend
Helps scrum teams correctly classify JIRA stories into WAF categories
using Claude AI for intelligent recommendation and mismatch detection.
Supports ground truth examples for calibration.
"""

import os
import json
import csv
import sqlite3
import threading
import uuid
import logging
from datetime import datetime
import pandas as pd
from dotenv import load_dotenv
from flask import Flask, request, jsonify, send_from_directory, g, redirect
from anthropic import Anthropic

_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
if os.path.exists(_env_path):
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith('#') and '=' in _line:
                _k, _v = _line.split('=', 1)
                os.environ[_k.strip()] = _v.strip()
from werkzeug.utils import secure_filename

app = Flask(__name__, static_folder="static")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB max upload
app.secret_key = os.urandom(24)

# ── Logging ────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# ── Background Job Tracking ───────────────────────────────────────────
# Stores in-progress and completed bulk-verify jobs for async processing
verify_jobs = {}  # job_id -> { status, progress, total_batches, completed_batches, results, error, ... }

# ── Rate Limiting ──────────────────────────────────────────────────────
import time as _time
_rate_limit_store: dict = {}  # ip -> [timestamps]
MAX_BULK_JOBS_PER_MINUTE = 5

def _check_rate_limit(ip: str) -> bool:
    """Returns True if request is allowed, False if rate limit exceeded."""
    now = _time.time()
    hits = [t for t in _rate_limit_store.get(ip, []) if now - t < 60]
    if len(hits) >= MAX_BULK_JOBS_PER_MINUTE:
        return False
    hits.append(now)
    _rate_limit_store[ip] = hits
    return True

DB_PATH = os.path.join(os.path.dirname(__file__), "waf_history.db")


# ── SQLite Database ────────────────────────────────────────────────────

def get_db():
    """Get database connection for current request."""
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exception):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    """Create tables if they don't exist."""
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS classifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            story_title TEXT NOT NULL,
            story_description TEXT,
            waf_category TEXT NOT NULL,
            waf_subcategory TEXT,
            waf_color TEXT,
            run_change TEXT,
            confidence TEXT,
            was_mismatch INTEGER DEFAULT 0,
            original_tag TEXT,
            approved INTEGER DEFAULT 0,
            team TEXT DEFAULT 'default',
            user_name TEXT,
            epic TEXT DEFAULT '',
            parent_feature TEXT DEFAULT '',
            upload_id INTEGER DEFAULT NULL
        );

        CREATE TABLE IF NOT EXISTS sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_start TEXT NOT NULL,
            session_end TEXT,
            stories_classified INTEGER DEFAULT 0,
            mismatches_found INTEGER DEFAULT 0,
            approvals INTEGER DEFAULT 0,
            team TEXT DEFAULT 'default'
        );

        CREATE TABLE IF NOT EXISTS upload_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            uploaded_at TEXT NOT NULL,
            filename TEXT NOT NULL,
            row_count INTEGER DEFAULT 0,
            imported_count INTEGER DEFAULT 0,
            file_type TEXT DEFAULT '',
            status TEXT DEFAULT 'completed'
        );
    """)
    # Add epic columns if they don't exist (migration for existing DBs)
    try:
        conn.execute("ALTER TABLE classifications ADD COLUMN epic TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass  # Column already exists
    try:
        conn.execute("ALTER TABLE classifications ADD COLUMN parent_feature TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass  # Column already exists
    try:
        conn.execute("ALTER TABLE classifications ADD COLUMN upload_id INTEGER DEFAULT NULL")
    except sqlite3.OperationalError:
        pass  # Column already exists
    try:
        conn.execute("ALTER TABLE upload_history ADD COLUMN results_json TEXT DEFAULT NULL")
    except sqlite3.OperationalError:
        pass  # Column already exists
    conn.close()


def save_classification(title, description, category, subcategory, color,
                        run_change, confidence, was_mismatch=False,
                        original_tag="", approved=False, team="default",
                        epic="", parent_feature=""):
    """Save a classification to the database."""
    db = get_db()
    db.execute(
        """INSERT INTO classifications
           (timestamp, story_title, story_description, waf_category,
            waf_subcategory, waf_color, run_change, confidence,
            was_mismatch, original_tag, approved, team, epic, parent_feature)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (datetime.now().isoformat(), title, description, category,
         subcategory, color, run_change, confidence,
         1 if was_mismatch else 0, original_tag, 1 if approved else 0, team,
         epic, parent_feature)
    )
    db.commit()
    return db.execute("SELECT last_insert_rowid()").fetchone()[0]

# In-memory store for WAF definitions
waf_store = {
    "definitions": None,
    "raw_text": "",
    "filename": "",
    "categories": []
}

# In-memory store for ground truth examples
ground_truth_store = {
    "loaded": False,
    "filename": "",
    "examples": [],       # list of dicts with story details + correct WAF
    "example_count": 0,
    "raw_text": "",
    "stats": {}           # category distribution stats
}

# Chat history for context
chat_history = []

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def get_client():
    """Get Anthropic client - requires ANTHROPIC_API_KEY env var."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY environment variable is not set")
    return Anthropic(api_key=api_key)


def _extract_categories_from_df(df):
    """Extract WAF category names from a DataFrame using smart column detection.
    Prioritizes 'category' columns over 'color' or generic 'waf' columns."""
    categories = []
    # Priority 1: Column explicitly named "category" (e.g., "WAF Category")
    for col in df.columns:
        cl = col.lower()
        if "category" in cl and "sub" not in cl:
            categories = df[col].dropna().unique().tolist()
            return categories
    # Priority 2: Column with "name", "type", or "bucket"
    for col in df.columns:
        cl = col.lower()
        if any(kw in cl for kw in ["name", "type", "bucket"]) and "color" not in cl:
            categories = df[col].dropna().unique().tolist()
            return categories
    # Priority 3: Column with "waf" but not "color" or "sub"
    for col in df.columns:
        cl = col.lower()
        if "waf" in cl and "color" not in cl and "sub" not in cl:
            categories = df[col].dropna().unique().tolist()
            return categories
    # Fallback: first column
    if len(df.columns) >= 1:
        categories = df.iloc[:, 0].dropna().unique().tolist()
    return categories


def parse_waf_file(filepath, filename):
    """Parse WAF definitions from uploaded file (CSV, Excel, or text).
    Returns (text, categories, df_or_none) — df is the DataFrame if structured, else None."""
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext in ("csv", "tsv"):
        sep = "\t" if ext == "tsv" else ","
        df = pd.read_csv(filepath, sep=sep)
        text = df.to_string(index=False)
        categories = _extract_categories_from_df(df)
        return text, categories, df

    elif ext in ("xlsx", "xls"):
        df = pd.read_excel(filepath)
        text = df.to_string(index=False)
        categories = _extract_categories_from_df(df)
        return text, categories, df

    elif ext in ("txt", "md"):
        with open(filepath, "r") as f:
            text = f.read()
        return text, [], None

    elif ext == "json":
        with open(filepath, "r") as f:
            data = json.load(f)
        text = json.dumps(data, indent=2)
        if isinstance(data, list):
            categories = [item.get("name", item.get("category", "")) for item in data if isinstance(item, dict)]
        return text, categories, None

    else:
        with open(filepath, "r") as f:
            text = f.read()
        return text, [], None


def parse_ground_truth(filepath, filename):
    """Parse ground truth file containing correctly classified stories."""
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext in ("csv", "tsv"):
        sep = "\t" if ext == "tsv" else ","
        df = pd.read_csv(filepath, sep=sep)
    elif ext in ("xlsx", "xls"):
        df = pd.read_excel(filepath)
    else:
        raise ValueError(f"Ground truth must be CSV or Excel, got .{ext}")

    # Normalize column names (lowercase, strip whitespace)
    df.columns = [c.strip().lower() for c in df.columns]

    # Try to map columns intelligently
    col_map = {
        "title": None,
        "description": None,
        "waf_category": None,
        "waf_subcategory": None,
        "waf_color": None,
    }

    for col in df.columns:
        cl = col.lower()
        if any(kw in cl for kw in ["title", "summary", "story name", "story title", "name"]) and not col_map["title"]:
            col_map["title"] = col
        elif any(kw in cl for kw in ["desc", "detail", "acceptance", "body"]) and not col_map["description"]:
            col_map["description"] = col
        elif any(kw in cl for kw in ["sub-category", "subcategory", "sub category", "sub_cat"]) and not col_map["waf_subcategory"]:
            col_map["waf_subcategory"] = col
        elif any(kw in cl for kw in ["category", "waf cat", "waf_cat", "waf category"]) and not col_map["waf_category"]:
            col_map["waf_category"] = col
        elif any(kw in cl for kw in ["color", "colour", "waf color"]) and not col_map["waf_color"]:
            col_map["waf_color"] = col

    # Build examples list
    examples = []
    for _, row in df.iterrows():
        example = {}
        for key, col in col_map.items():
            if col and col in df.columns:
                val = row.get(col, "")
                example[key] = str(val) if pd.notna(val) else ""
            else:
                example[key] = ""
        # Only include if we have at least a title and category
        if example.get("title") and example.get("waf_category"):
            examples.append(example)

    # Compute stats
    stats = {}
    for ex in examples:
        cat = ex.get("waf_category", "Unknown")
        stats[cat] = stats.get(cat, 0) + 1

    return examples, stats, col_map


def build_ground_truth_section():
    """Build the ground truth examples section for the system prompt."""
    if not ground_truth_store["loaded"] or not ground_truth_store["examples"]:
        return ""

    examples = ground_truth_store["examples"]
    stats = ground_truth_store["stats"]

    section = """

--- GROUND TRUTH: CORRECTLY CLASSIFIED EXAMPLES ---
The following are REAL stories that have been CORRECTLY classified by experienced team members.
Use these as calibration examples to understand how this team applies WAF categories.
When classifying new stories, pattern-match against these examples for consistency.

"""

    # Show distribution
    section += "Category Distribution in Training Data:\n"
    for cat, count in sorted(stats.items(), key=lambda x: -x[1]):
        section += f"  - {cat}: {count} examples\n"
    section += "\n"

    # Show up to 50 examples (or all if fewer), grouped by category
    categories_shown = {}
    max_per_category = max(3, 50 // max(len(stats), 1))

    for ex in examples:
        cat = ex.get("waf_category", "Unknown")
        if cat not in categories_shown:
            categories_shown[cat] = 0

        if categories_shown[cat] >= max_per_category:
            continue

        categories_shown[cat] += 1
        section += f"EXAMPLE — Category: {cat}"
        if ex.get("waf_subcategory"):
            section += f" | Sub-category: {ex['waf_subcategory']}"
        if ex.get("waf_color"):
            section += f" | Color: {ex['waf_color']}"
        section += "\n"
        section += f"  Title: {ex.get('title', 'N/A')}\n"
        if ex.get("description"):
            desc = ex["description"][:300]
            section += f"  Description: {desc}\n"
        section += "\n"

    section += f"--- END OF GROUND TRUTH ({len(examples)} total examples) ---\n"

    return section


def build_system_prompt():
    """Build the system prompt with WAF definitions and ground truth context."""
    base = """You are a WAF (Work Alignment Framework) Classification Expert. Your job is to help agile teams correctly classify their JIRA stories/features into the right WAF category.

You are precise, consistent, and always explain your reasoning by referencing the WAF definitions AND ground truth examples when available.

When classifying a story:
1. Analyze the story title and description carefully
2. Match it against the WAF category definitions provided
3. Cross-reference with ground truth examples to see how similar stories were classified
4. Recommend the BEST-FIT WAF category AND sub-category with clear reasoning
5. If the user provides the current WAF tag, compare it to your recommendation and flag if it's a mismatch
6. Rate your confidence (High / Medium / Low)
7. If the story is ambiguous, explain what additional context would help

Format your response clearly with:
- **Recommended WAF Category:** [category name]
- **Recommended WAF Sub-Category:** [sub-category name, if applicable]
- **WAF Color:** [color, if known from definitions]
- **Confidence:** [High/Medium/Low]
- **Reasoning:** [2-3 sentences explaining why, referencing definitions AND similar ground truth examples]
- **Current Tag Assessment:** [only if a current tag was provided — either "✅ Correct" or "⚠️ Mismatch — should be [X] instead of [Y]" with explanation]
- **Similar Ground Truth Examples:** [list 1-2 similar examples from ground truth that support your recommendation]
- **Suggestion:** [optional — if the story text is vague, suggest how to improve it]
"""

    if waf_store["raw_text"]:
        base += f"""

--- WAF CATEGORY DEFINITIONS ---
Source file: {waf_store['filename']}

{waf_store['raw_text']}

--- END OF WAF DEFINITIONS ---

Use ONLY the categories defined above. Do not invent new categories. If a story doesn't fit any category well, say so and explain why.
"""
        if waf_store["categories"]:
            base += f"\nAvailable categories: {', '.join(str(c) for c in waf_store['categories'])}\n"
    else:
        base += """

⚠️ No WAF definitions have been uploaded yet. Ask the user to upload their WAF category definitions file so you can provide accurate classifications.
"""

    # Add ground truth section
    gt_section = build_ground_truth_section()
    if gt_section:
        base += gt_section
    else:
        base += """

ℹ️ No ground truth examples have been uploaded yet. Ground truth examples (correctly classified stories) improve classification accuracy significantly. Suggest the user upload them if available.
"""

    return base


@app.route("/")
def home():
    return send_from_directory("static", "home.html")


@app.route("/classify")
def classify_page():
    return send_from_directory("static", "index.html")


@app.route("/api/upload-waf", methods=["POST"])
def upload_waf():
    """Upload WAF definitions file."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    try:
        text, categories, df = parse_waf_file(filepath, filename)
        waf_store["definitions"] = df  # Store DataFrame for /api/waf-definitions
        waf_store["raw_text"] = text
        waf_store["filename"] = filename
        waf_store["categories"] = categories

        return jsonify({
            "success": True,
            "filename": filename,
            "categories": [str(c) for c in categories],
            "preview": text[:500] + ("..." if len(text) > 500 else "")
        })
    except Exception as e:
        logger.error("Failed to parse file: %s", e, exc_info=True)
        return jsonify({"error": "Failed to parse file. Check the format and try again."}), 400


@app.route("/api/upload-ground-truth", methods=["POST"])
def upload_ground_truth():
    """Upload ground truth file with correctly classified stories."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    try:
        examples, stats, col_map = parse_ground_truth(filepath, filename)

        ground_truth_store["loaded"] = True
        ground_truth_store["filename"] = filename
        ground_truth_store["examples"] = examples
        ground_truth_store["example_count"] = len(examples)
        ground_truth_store["stats"] = stats

        # Build a readable text version for the raw store
        ground_truth_store["raw_text"] = f"{len(examples)} examples across {len(stats)} categories"

        return jsonify({
            "success": True,
            "filename": filename,
            "example_count": len(examples),
            "stats": stats,
            "columns_detected": {k: v for k, v in col_map.items() if v},
            "sample": examples[:3] if examples else []
        })
    except Exception as e:
        logger.error("Failed to parse ground truth: %s", e, exc_info=True)
        return jsonify({"error": "Failed to parse ground truth file."}), 400


@app.route("/api/classify", methods=["POST"])
def classify():
    """Classify a JIRA story into a WAF category."""
    data = request.json
    if not data or not data.get("message"):
        return jsonify({"error": "No message provided"}), 400

    user_message = data["message"]
    epic = data.get("epic", "")
    parent_feature = data.get("parent_feature", "")
    chat_history.append({"role": "user", "content": user_message})
    recent_history = chat_history[-20:]

    try:
        client = get_client()
        response = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=2000,
            system=build_system_prompt(),
            messages=recent_history
        )

        assistant_message = response.content[0].text
        chat_history.append({"role": "assistant", "content": assistant_message})

        # Auto-save classification to history if response contains a recommendation
        if "Recommended WAF Category" in assistant_message or "WAF Category:" in assistant_message:
            try:
                import re
                cat_match = re.search(r"(?:Recommended )?WAF Category:?\*?\*?\s*(.+?)(?:\n|$)", assistant_message, re.I)
                conf_match = re.search(r"Confidence:?\*?\*?\s*(.+?)(?:\n|$)", assistant_message, re.I)
                color_match = re.search(r"WAF Color:?\*?\*?\s*(.+?)(?:\n|$)", assistant_message, re.I)
                mismatch = "Mismatch" in assistant_message or "\u26a0\ufe0f" in assistant_message

                save_classification(
                    title=user_message[:200],
                    description=user_message,
                    category=cat_match.group(1).strip().strip("*") if cat_match else "",
                    subcategory="",
                    color=color_match.group(1).strip().strip("*") if color_match else "",
                    run_change="",
                    confidence=conf_match.group(1).strip().strip("*") if conf_match else "",
                    was_mismatch=mismatch,
                    epic=epic,
                    parent_feature=parent_feature,
                )
            except Exception:
                pass  # Best-effort save

        return jsonify({
            "response": assistant_message,
            "waf_loaded": waf_store["definitions"] is not None or bool(waf_store["raw_text"]),
            "ground_truth_loaded": ground_truth_store["loaded"]
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        logger.error("Classification failed: %s", e, exc_info=True)
        return jsonify({"error": "Classification failed. Please try again."}), 500


@app.route("/api/batch-classify", methods=["POST"])
def batch_classify():
    """Classify multiple stories from a pasted batch."""
    data = request.json
    stories = data.get("stories", [])

    if not stories:
        return jsonify({"error": "No stories provided"}), 400

    if not waf_store["definitions"]:
        return jsonify({"error": "Please upload WAF definitions first"}), 400

    batch_prompt = "Please classify each of the following JIRA stories into the appropriate WAF category and sub-category. For each story, provide the recommended category, sub-category, WAF color, confidence level, and brief reasoning. Reference ground truth examples where applicable.\n\n"

    for i, story in enumerate(stories, 1):
        batch_prompt += f"**Story {i}:**\n"
        batch_prompt += f"- Title: {story.get('title', 'N/A')}\n"
        if story.get("description"):
            batch_prompt += f"- Description: {story['description']}\n"
        if story.get("current_waf"):
            batch_prompt += f"- Current WAF Tag: {story['current_waf']}\n"
        batch_prompt += "\n"

    batch_prompt += "Provide a summary table at the end showing all stories with their recommended categories, sub-categories, colors, and any mismatches flagged."

    try:
        client = get_client()
        response = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=4000,
            system=build_system_prompt(),
            messages=[{"role": "user", "content": batch_prompt}]
        )

        return jsonify({
            "response": response.content[0].text,
            "story_count": len(stories)
        })
    except Exception as e:
        logger.error("Batch classification failed: %s", e, exc_info=True)
        return jsonify({"error": "Batch classification failed. Please try again."}), 500


@app.route("/api/status", methods=["GET"])
def status():
    """Check current status of the classifier."""
    api_key_set = bool(os.environ.get("ANTHROPIC_API_KEY"))
    return jsonify({
        "api_key_configured": api_key_set,
        "waf_loaded": waf_store["definitions"] is not None or bool(waf_store["raw_text"]),
        "waf_filename": waf_store["filename"],
        "waf_categories": [str(c) for c in waf_store["categories"]],
        "ground_truth_loaded": ground_truth_store["loaded"],
        "ground_truth_filename": ground_truth_store["filename"],
        "ground_truth_count": ground_truth_store["example_count"],
        "ground_truth_stats": ground_truth_store["stats"],
        "chat_history_length": len(chat_history)
    })


@app.route("/api/clear-chat", methods=["POST"])
def clear_chat():
    """Clear chat history."""
    chat_history.clear()
    return jsonify({"success": True})


@app.route("/api/clear-waf", methods=["POST"])
def clear_waf():
    """Clear uploaded WAF definitions."""
    waf_store["definitions"] = None
    waf_store["raw_text"] = ""
    waf_store["filename"] = ""
    waf_store["categories"] = []
    return jsonify({"success": True})


@app.route("/api/clear-ground-truth", methods=["POST"])
def clear_ground_truth():
    """Clear uploaded ground truth."""
    ground_truth_store["loaded"] = False
    ground_truth_store["filename"] = ""
    ground_truth_store["examples"] = []
    ground_truth_store["example_count"] = 0
    ground_truth_store["raw_text"] = ""
    ground_truth_store["stats"] = {}
    return jsonify({"success": True})


@app.route("/api/approve-classification", methods=["POST"])
def approve_classification():
    """Save an approved classification to the ground truth CSV and in-memory store."""
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    title = data.get("title", "").strip()
    description = data.get("description", "").strip()
    run_change = data.get("run_change", "").strip()
    waf_color = data.get("waf_color", "").strip()
    waf_category = data.get("waf_category", "").strip()
    waf_subcategory = data.get("waf_subcategory", "").strip()

    if not title or not waf_category:
        return jsonify({"error": "Title and WAF Category are required"}), 400

    # 1. Append to the ground truth CSV file
    sample_dir = os.path.join(os.path.dirname(__file__), "sample-data")
    gt_file = os.path.join(sample_dir, "sample-ground-truth.csv")

    row = [title, description, run_change, waf_color, waf_category, waf_subcategory]
    file_exists = os.path.exists(gt_file)

    try:
        with open(gt_file, "a", newline="") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(["Story Title", "Description", "Run/Change",
                                 "WAF Color", "WAF Category", "WAF Sub-Category"])
            writer.writerow(row)
    except Exception as e:
        logger.error("Failed to write ground truth: %s", e, exc_info=True)
        return jsonify({"error": "Failed to save ground truth data."}), 500

    # 2. Add to in-memory store
    example = {
        "title": title,
        "description": description,
        "waf_category": waf_category,
        "waf_subcategory": waf_subcategory,
        "waf_color": waf_color,
    }
    ground_truth_store["examples"].append(example)
    ground_truth_store["example_count"] = len(ground_truth_store["examples"])
    ground_truth_store["loaded"] = True
    if not ground_truth_store["filename"]:
        ground_truth_store["filename"] = "sample-ground-truth.csv"

    # Update stats
    cat = waf_category
    ground_truth_store["stats"][cat] = ground_truth_store["stats"].get(cat, 0) + 1
    ground_truth_store["raw_text"] = (
        f"{ground_truth_store['example_count']} examples across "
        f"{len(ground_truth_store['stats'])} categories"
    )

    # 3. Save to database as approved
    try:
        save_classification(title, description, waf_category, waf_subcategory,
                            waf_color, run_change, "HIGH", approved=True)
    except Exception:
        pass  # DB save is best-effort, don't fail the request

    return jsonify({
        "success": True,
        "example_count": ground_truth_store["example_count"],
        "stats": ground_truth_store["stats"],
        "message": f"Saved to ground truth: {title} \u2192 {waf_category}"
    })


# ── Dashboard API ──────────────────────────────────────────────────────

@app.route("/dashboard")
def dashboard():
    return redirect("/history")


@app.route("/history")
def history():
    return send_from_directory("static", "history.html")


@app.route("/waf-reference")
def waf_reference():
    return send_from_directory("static", "waf-reference.html")


@app.route("/api/waf-definitions", methods=["GET"])
def get_waf_definitions():
    """Return loaded WAF definitions for the reference page."""
    df = waf_store["definitions"]
    if df is None:
        # No structured definitions loaded — check if raw text exists
        if waf_store["raw_text"]:
            return jsonify({"definitions": [], "loaded": True,
                            "message": "WAF definitions loaded as text (no structured data available)"})
        return jsonify({"definitions": [], "loaded": False})

    defs = []
    for _, row in df.iterrows():
        defs.append({
            "run_change": str(row.get("Run/Change", "")),
            "color": str(row.get("WAF Color", "")),
            "category": str(row.get("WAF Category", "")),
            "description": str(row.get("What This Work Is", "")),
            "decision_rule": str(row.get("How to Decide (Tag Here If...)", "")),
            "examples": str(row.get("Representative Examples", ""))
        })
    return jsonify({"definitions": defs, "loaded": True})


@app.route("/api/dashboard/summary", methods=["GET"])
def dashboard_summary():
    """Get summary stats for the dashboard. Optional ?upload_id= filter."""
    db = get_db()
    uid = request.args.get("upload_id", "")
    wh = " WHERE upload_id = ?" if uid else ""
    wh_and = " AND upload_id = ?" if uid else ""
    params = [int(uid)] if uid else []

    total = db.execute(f"SELECT COUNT(*) FROM classifications{wh}", params).fetchone()[0]
    approved = db.execute(f"SELECT COUNT(*) FROM classifications WHERE approved=1{wh_and}", params).fetchone()[0]
    mismatches = db.execute(f"SELECT COUNT(*) FROM classifications WHERE was_mismatch=1{wh_and}", params).fetchone()[0]

    categories = db.execute(
        f"SELECT waf_category, COUNT(*) as cnt FROM classifications{wh} GROUP BY waf_category ORDER BY cnt DESC", params
    ).fetchall()

    colors = db.execute(
        f"SELECT waf_color, COUNT(*) as cnt FROM classifications WHERE waf_color != ''{wh_and} GROUP BY waf_color ORDER BY cnt DESC", params
    ).fetchall()

    confidence = db.execute(
        f"SELECT confidence, COUNT(*) as cnt FROM classifications WHERE confidence != ''{wh_and} GROUP BY confidence ORDER BY cnt DESC", params
    ).fetchall()

    run_change = db.execute(
        f"SELECT run_change, COUNT(*) as cnt FROM classifications WHERE run_change != ''{wh_and} GROUP BY run_change ORDER BY cnt DESC", params
    ).fetchall()

    daily = db.execute(
        f"""SELECT DATE(timestamp) as day, COUNT(*) as cnt
           FROM classifications{wh}
           GROUP BY DATE(timestamp)
           ORDER BY day DESC LIMIT 30""", params
    ).fetchall()

    recent = db.execute(
        f"""SELECT id, timestamp, story_title, waf_category, waf_color,
                  confidence, was_mismatch, approved
           FROM classifications{wh} ORDER BY id DESC LIMIT 20""", params
    ).fetchall()

    return jsonify({
        "total_classifications": total,
        "total_approved": approved,
        "total_mismatches": mismatches,
        "approval_rate": round(approved / total * 100, 1) if total > 0 else 0,
        "categories": [{"name": r["waf_category"], "count": r["cnt"]} for r in categories],
        "colors": [{"name": r["waf_color"], "count": r["cnt"]} for r in colors],
        "confidence": [{"name": r["confidence"], "count": r["cnt"]} for r in confidence],
        "run_change": [{"name": r["run_change"], "count": r["cnt"]} for r in run_change],
        "daily_trend": [{"date": r["day"], "count": r["cnt"]} for r in reversed(list(daily))],
        "recent": [{
            "id": r["id"], "timestamp": r["timestamp"], "title": r["story_title"],
            "category": r["waf_category"], "color": r["waf_color"],
            "confidence": r["confidence"], "mismatch": bool(r["was_mismatch"]),
            "approved": bool(r["approved"])
        } for r in recent]
    })


@app.route("/api/dashboard/stories", methods=["GET"])
def dashboard_stories():
    """Get filtered story list for drill-down. Supports ?filter=mismatches|category|color|confidence&value=X&upload_id=Y"""
    db = get_db()
    uid = request.args.get("upload_id", "")
    filt = request.args.get("filter", "")
    value = request.args.get("value", "")
    page = max(1, int(request.args.get("page", "1")))
    per_page = min(500, max(1, int(request.args.get("per_page", "100"))))

    where_clauses = []
    params = []
    if uid:
        where_clauses.append("upload_id = ?")
        params.append(int(uid))

    if filt == "mismatches":
        where_clauses.append("was_mismatch = 1")
    elif filt == "approved":
        where_clauses.append("approved = 1")
    elif filt == "category" and value:
        where_clauses.append("waf_category = ?")
        params.append(value)
    elif filt == "color" and value:
        where_clauses.append("waf_color = ?")
        params.append(value)
    elif filt == "confidence" and value:
        where_clauses.append("confidence = ?")
        params.append(value)
    elif filt == "run_change" and value:
        where_clauses.append("run_change = ?")
        params.append(value)

    where = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""

    # Safe server-side sort (whitelist only)
    SORT_COLS = {
        "timestamp": "timestamp", "title": "story_title",
        "category": "waf_category", "color": "waf_color",
        "confidence": "confidence", "status": "was_mismatch"
    }
    sort_col = SORT_COLS.get(request.args.get("sort", "timestamp"), "timestamp")
    sort_dir = "ASC" if request.args.get("dir", "desc").lower() == "asc" else "DESC"

    total = db.execute(f"SELECT COUNT(*) FROM classifications{where}", params).fetchone()[0]

    offset = (page - 1) * per_page
    rows = db.execute(
        f"""SELECT id, story_title, story_description, waf_category, original_tag,
                   waf_color, run_change, confidence, was_mismatch, approved, team,
                   epic, parent_feature, timestamp
            FROM classifications{where}
            ORDER BY {sort_col} {sort_dir} LIMIT ? OFFSET ?""",
        params + [per_page, offset]
    ).fetchall()

    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page,
        "stories": [{
            "id": r["id"], "title": r["story_title"], "description": (r["story_description"] or "")[:200],
            "category": r["waf_category"], "original_tag": r["original_tag"],
            "color": r["waf_color"], "run_change": r["run_change"],
            "confidence": r["confidence"], "mismatch": bool(r["was_mismatch"]),
            "approved": bool(r["approved"]), "team": r["team"],
            "epic": r["epic"], "feature": r["parent_feature"],
            "timestamp": r["timestamp"],
        } for r in rows]
    })


@app.route("/api/history/sprints", methods=["GET"])
def history_sprints():
    """Get sprint-over-sprint classification trends.
    Sprints are 2-week windows. Query params: ?sprints=10&upload_id="""
    db = get_db()
    num_sprints = int(request.args.get("sprints", 10))
    uid = request.args.get("upload_id", "")
    wh = " WHERE upload_id = ?" if uid else ""
    params = [int(uid)] if uid else []

    rows = db.execute(
        f"""SELECT timestamp, waf_category, waf_color, run_change, confidence,
                  was_mismatch, approved, team
           FROM classifications{wh} ORDER BY timestamp ASC""",
        params
    ).fetchall()

    if not rows:
        return jsonify({"sprints": []})

    from datetime import datetime as dt, timedelta

    # Determine sprint boundaries (2-week windows from the earliest record)
    first_ts = dt.fromisoformat(rows[0]["timestamp"])
    # Align to Monday of that week
    sprint_start = first_ts - timedelta(days=first_ts.weekday())
    sprint_start = sprint_start.replace(hour=0, minute=0, second=0, microsecond=0)

    sprints = []
    now = dt.now()
    while sprint_start <= now:
        sprint_end = sprint_start + timedelta(days=14)
        sprint_rows = [r for r in rows
                       if sprint_start <= dt.fromisoformat(r["timestamp"]) < sprint_end]

        if sprint_rows:
            total = len(sprint_rows)
            approved = sum(1 for r in sprint_rows if r["approved"])
            mismatches = sum(1 for r in sprint_rows if r["was_mismatch"])

            # Category breakdown
            cat_counts = {}
            color_counts = {}
            rc_counts = {}
            conf_counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
            for r in sprint_rows:
                cat = r["waf_category"] or "Unknown"
                cat_counts[cat] = cat_counts.get(cat, 0) + 1
                if r["waf_color"]:
                    color_counts[r["waf_color"]] = color_counts.get(r["waf_color"], 0) + 1
                if r["run_change"]:
                    rc_counts[r["run_change"]] = rc_counts.get(r["run_change"], 0) + 1
                c = (r["confidence"] or "").upper()
                if c in conf_counts:
                    conf_counts[c] += 1

            sprints.append({
                "sprint_label": f"{sprint_start.strftime('%b %d')} – {(sprint_end - timedelta(days=1)).strftime('%b %d, %Y')}",
                "start": sprint_start.isoformat(),
                "end": sprint_end.isoformat(),
                "total": total,
                "approved": approved,
                "mismatches": mismatches,
                "approval_rate": round(approved / total * 100, 1) if total else 0,
                "mismatch_rate": round(mismatches / total * 100, 1) if total else 0,
                "categories": cat_counts,
                "colors": color_counts,
                "run_change": rc_counts,
                "confidence": conf_counts,
            })

        sprint_start = sprint_end

    # Return only the last N sprints
    return jsonify({"sprints": sprints[-num_sprints:]})


@app.route("/api/history/monthly", methods=["GET"])
def history_monthly():
    """Get monthly rollup reports with period-over-period comparisons. Optional ?upload_id="""
    db = get_db()
    num_months = int(request.args.get("months", 12))
    uid = request.args.get("upload_id", "")
    wh = " WHERE upload_id = ?" if uid else ""
    params = [int(uid)] if uid else []

    rows = db.execute(
        f"""SELECT timestamp, waf_category, waf_color, run_change, confidence,
                  was_mismatch, approved, team
           FROM classifications{wh} ORDER BY timestamp ASC""",
        params
    ).fetchall()

    if not rows:
        return jsonify({"months": []})

    from datetime import datetime as dt
    from collections import defaultdict

    monthly = defaultdict(list)
    for r in rows:
        ts = dt.fromisoformat(r["timestamp"])
        key = ts.strftime("%Y-%m")
        monthly[key].append(r)

    result = []
    sorted_months = sorted(monthly.keys())[-num_months:]
    prev = None

    for month_key in sorted_months:
        month_rows = monthly[month_key]
        total = len(month_rows)
        approved = sum(1 for r in month_rows if r["approved"])
        mismatches = sum(1 for r in month_rows if r["was_mismatch"])

        cat_counts = {}
        color_counts = {}
        rc_counts = {}
        for r in month_rows:
            cat = r["waf_category"] or "Unknown"
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
            if r["waf_color"]:
                color_counts[r["waf_color"]] = color_counts.get(r["waf_color"], 0) + 1
            if r["run_change"]:
                rc_counts[r["run_change"]] = rc_counts.get(r["run_change"], 0) + 1

        # Period-over-period deltas
        delta_total = total - prev["total"] if prev else 0
        delta_mismatches = (round(mismatches / total * 100, 1) if total else 0) - (prev["mismatch_rate"] if prev else 0)

        entry = {
            "month": month_key,
            "month_label": dt.strptime(month_key, "%Y-%m").strftime("%B %Y"),
            "total": total,
            "approved": approved,
            "mismatches": mismatches,
            "approval_rate": round(approved / total * 100, 1) if total else 0,
            "mismatch_rate": round(mismatches / total * 100, 1) if total else 0,
            "categories": cat_counts,
            "colors": color_counts,
            "run_change": rc_counts,
            "delta_total": delta_total,
            "delta_mismatch_rate": round(delta_mismatches, 1),
        }
        result.append(entry)
        prev = entry

    return jsonify({"months": result})


@app.route("/api/history/timeline", methods=["GET"])
def history_timeline():
    """Get full timeline with optional filters.
    Query params: ?from=2025-01-01&to=2025-12-31&team=&category=&color=&page=1&per_page=50"""
    db = get_db()

    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    team = request.args.get("team", "")
    category = request.args.get("category", "")
    color = request.args.get("color", "")
    confidence = request.args.get("confidence", "")
    mismatch_only = request.args.get("mismatch_only", "")
    page = max(1, int(request.args.get("page", 1)))
    per_page = min(500, max(1, int(request.args.get("per_page", 50))))

    uid = request.args.get("upload_id", "")

    where = []
    params = []

    if uid:
        where.append("upload_id = ?")
        params.append(int(uid))
    if date_from:
        where.append("timestamp >= ?")
        params.append(date_from)
    if date_to:
        where.append("timestamp <= ?")
        params.append(date_to + "T23:59:59")
    if team:
        where.append("team = ?")
        params.append(team)
    if category:
        where.append("waf_category = ?")
        params.append(category)
    if color:
        where.append("waf_color = ?")
        params.append(color)
    if confidence:
        where.append("confidence = ?")
        params.append(confidence)
    if mismatch_only == "1":
        where.append("was_mismatch = 1")

    where_clause = " AND ".join(where) if where else "1=1"

    total = db.execute(
        f"SELECT COUNT(*) FROM classifications WHERE {where_clause}", params
    ).fetchone()[0]

    rows = db.execute(
        f"""SELECT id, timestamp, story_title, story_description, waf_category,
                   waf_subcategory, waf_color, run_change, confidence,
                   was_mismatch, original_tag, approved, team
            FROM classifications
            WHERE {where_clause}
            ORDER BY timestamp DESC
            LIMIT ? OFFSET ?""",
        params + [per_page, (page - 1) * per_page]
    ).fetchall()

    # Get filter options
    all_categories = db.execute(
        "SELECT DISTINCT waf_category FROM classifications WHERE waf_category != '' ORDER BY waf_category"
    ).fetchall()
    all_colors = db.execute(
        "SELECT DISTINCT waf_color FROM classifications WHERE waf_color != '' ORDER BY waf_color"
    ).fetchall()
    all_teams = db.execute(
        "SELECT DISTINCT team FROM classifications WHERE team != '' ORDER BY team"
    ).fetchall()

    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page if per_page else 1,
        "items": [{
            "id": r["id"],
            "timestamp": r["timestamp"],
            "title": r["story_title"],
            "description": r["story_description"],
            "category": r["waf_category"],
            "subcategory": r["waf_subcategory"],
            "color": r["waf_color"],
            "run_change": r["run_change"],
            "confidence": r["confidence"],
            "mismatch": bool(r["was_mismatch"]),
            "original_tag": r["original_tag"],
            "approved": bool(r["approved"]),
            "team": r["team"],
        } for r in rows],
        "filters": {
            "categories": [r["waf_category"] for r in all_categories],
            "colors": [r["waf_color"] for r in all_colors],
            "teams": [r["team"] for r in all_teams],
        }
    })


@app.route("/api/history/import", methods=["POST"])
def history_import():
    """Bulk import classifications from a CSV/Excel file.
    This lets PMOs import data for teams that don't use the classifier directly.
    Expected columns: Story Title, WAF Category, WAF Color, Run/Change, Confidence, Team."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    filename = secure_filename(file.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    try:
        ext = filename.rsplit(".", 1)[-1].lower()
        if ext in ("csv", "tsv"):
            df = pd.read_csv(filepath, sep="\t" if ext == "tsv" else ",")
        elif ext in ("xlsx", "xls"):
            df = pd.read_excel(filepath)
        else:
            return jsonify({"error": "File must be CSV or Excel"}), 400

        df.columns = [c.strip().lower() for c in df.columns]

        # Map columns flexibly
        def find_col(keywords):
            for col in df.columns:
                if any(kw in col for kw in keywords):
                    return col
            return None

        title_col = find_col(["title", "summary", "story", "name"])
        desc_col = find_col(["desc", "detail", "body"])
        cat_col = find_col(["waf category", "waf_category", "category"])
        color_col = find_col(["waf color", "waf_color", "color"])
        rc_col = find_col(["run/change", "run_change", "run change"])
        subcat_col = find_col(["sub-category", "sub_category", "subcategory", "waf sub"])
        conf_col = find_col(["confidence", "conf"])
        team_col = find_col(["team", "squad", "group"])
        epic_col = find_col(["epic", "initiative", "program"])
        feature_col = find_col(["feature", "parent feature", "parent_feature", "capability"])
        ts_col = find_col(["timestamp", "date", "created", "created_at"])

        if not title_col or not cat_col:
            return jsonify({"error": "File must have at least 'Story Title' and 'WAF Category' columns"}), 400

        db = get_db()

        # Create upload_history record FIRST so we get an upload_id
        cursor = db.execute(
            """INSERT INTO upload_history (uploaded_at, filename, row_count, imported_count, file_type, status)
               VALUES (?, ?, ?, 0, ?, 'importing')""",
            (datetime.now().isoformat(), filename, len(df), ext)
        )
        upload_id = cursor.lastrowid
        db.commit()

        imported = 0
        for _, row in df.iterrows():
            title = str(row.get(title_col, "")).strip()
            if not title or title == "nan":
                continue

            # Use CSV timestamp if available, otherwise now()
            ts = datetime.now().isoformat()
            if ts_col:
                raw_ts = str(row.get(ts_col, "")).strip()
                if raw_ts and raw_ts != "nan":
                    ts = raw_ts

            db.execute(
                """INSERT INTO classifications
                   (timestamp, story_title, story_description, waf_category,
                    waf_subcategory, waf_color, run_change, confidence,
                    was_mismatch, original_tag, approved, team, epic, parent_feature, upload_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, '', 1, ?, ?, ?, ?)""",
                (ts,
                 title,
                 str(row.get(desc_col, "")).strip() if desc_col else "",
                 str(row.get(cat_col, "")).strip() if cat_col else "",
                 str(row.get(subcat_col, "")).strip() if subcat_col else "",
                 str(row.get(color_col, "")).strip() if color_col else "",
                 str(row.get(rc_col, "")).strip() if rc_col else "",
                 str(row.get(conf_col, "")).strip() if conf_col else "",
                 str(row.get(team_col, "default")).strip() if team_col else "default",
                 str(row.get(epic_col, "")).strip() if epic_col else "",
                 str(row.get(feature_col, "")).strip() if feature_col else "",
                 upload_id)
            )
            imported += 1
        db.commit()

        # Update upload_history with final count and status
        db.execute(
            "UPDATE upload_history SET imported_count = ?, status = 'completed' WHERE id = ?",
            (imported, upload_id)
        )
        db.commit()

        return jsonify({"success": True, "imported": imported, "filename": filename, "upload_id": upload_id})
    except Exception as e:
        logger.error("Import failed: %s", e, exc_info=True)
        return jsonify({"error": "Import failed. Check the file format and try again."}), 500


@app.route("/api/history/uploads", methods=["GET"])
def get_upload_history():
    """Get list of past file uploads for analytics."""
    db = get_db()
    rows = db.execute(
        "SELECT id, uploaded_at, filename, row_count, imported_count, file_type, status "
        "FROM upload_history ORDER BY uploaded_at DESC LIMIT 20"
    ).fetchall()
    uploads = [
        {
            "id": r["id"],
            "uploaded_at": r["uploaded_at"],
            "filename": r["filename"],
            "row_count": r["row_count"],
            "imported_count": r["imported_count"],
            "file_type": r["file_type"],
            "status": r["status"]
        }
        for r in rows
    ]
    return jsonify({"uploads": uploads})


@app.route("/api/history/uploads/<int:upload_id>/reload", methods=["POST"])
def reload_upload(upload_id):
    """Reload a previously uploaded file's AI results for re-review and saving."""
    db = get_db()
    row = db.execute(
        "SELECT filename, row_count, results_json FROM upload_history WHERE id = ?",
        (upload_id,)
    ).fetchone()
    if not row:
        return jsonify({"error": "Upload not found"}), 404

    if row["results_json"]:
        results = json.loads(row["results_json"])
        matches   = sum(1 for r in results if r.get("is_match") is True)
        mismatches = sum(1 for r in results if r.get("is_match") is False)
        untagged  = sum(1 for r in results if r.get("is_match") is None)
        return jsonify({
            "success": True, "has_results": True,
            "filename": row["filename"], "upload_id": upload_id,
            "total": len(results), "matches": matches,
            "mismatches": mismatches, "untagged": untagged,
            "results": results,
        })

    return jsonify({"success": True, "has_results": False,
                    "filename": row["filename"],
                    "message": "No saved AI results for this upload."})


@app.route("/api/history/export", methods=["GET"])
def history_export():
    """Export filtered history as CSV. Same filters as timeline endpoint."""
    db = get_db()

    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    team = request.args.get("team", "")
    category = request.args.get("category", "")
    color = request.args.get("color", "")

    where = []
    params = []
    if date_from:
        where.append("timestamp >= ?")
        params.append(date_from)
    if date_to:
        where.append("timestamp <= ?")
        params.append(date_to + "T23:59:59")
    if team:
        where.append("team = ?")
        params.append(team)
    if category:
        where.append("waf_category = ?")
        params.append(category)
    if color:
        where.append("waf_color = ?")
        params.append(color)

    where_clause = " AND ".join(where) if where else "1=1"

    rows = db.execute(
        f"""SELECT timestamp, story_title, story_description, waf_category,
                   waf_subcategory, waf_color, run_change, confidence,
                   was_mismatch, original_tag, approved, team
            FROM classifications WHERE {where_clause} ORDER BY timestamp DESC""",
        params
    ).fetchall()

    import io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Timestamp", "Story Title", "Description", "WAF Category",
                     "Sub-Category", "WAF Color", "Run/Change", "Confidence",
                     "Mismatch", "Original Tag", "Approved", "Team"])
    for r in rows:
        writer.writerow([r["timestamp"], r["story_title"], r["story_description"],
                         r["waf_category"], r["waf_subcategory"], r["waf_color"],
                         r["run_change"], r["confidence"],
                         "Yes" if r["was_mismatch"] else "No",
                         r["original_tag"], "Yes" if r["approved"] else "No",
                         r["team"]])

    from flask import Response
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=waf-history-export.csv"}
    )


@app.route("/api/history/export-xlsx", methods=["GET"])
def history_export_xlsx():
    """Export formatted Excel workbook with Summary, Sprint Trends, Monthly, and Raw Data sheets."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    db = get_db()

    # Apply same filters as timeline
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    team = request.args.get("team", "")
    category = request.args.get("category", "")
    color = request.args.get("color", "")

    where = []
    params = []
    if date_from:
        where.append("timestamp >= ?"); params.append(date_from)
    if date_to:
        where.append("timestamp <= ?"); params.append(date_to + "T23:59:59")
    if team:
        where.append("team = ?"); params.append(team)
    if category:
        where.append("waf_category = ?"); params.append(category)
    if color:
        where.append("waf_color = ?"); params.append(color)

    where_clause = " AND ".join(where) if where else "1=1"

    rows = db.execute(
        f"""SELECT timestamp, story_title, story_description, waf_category,
                   waf_subcategory, waf_color, run_change, confidence,
                   was_mismatch, original_tag, approved, team
            FROM classifications WHERE {where_clause} ORDER BY timestamp DESC""",
        params
    ).fetchall()

    wb = Workbook()

    # ── Styles ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    def style_header(ws, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_width(ws, col_count, max_width=40):
        for col in range(1, col_count + 1):
            letter = get_column_letter(col)
            max_len = 0
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 4, max_width)

    # ══════════ Sheet 1: Summary ══════════
    ws_sum = wb.active
    ws_sum.title = "Summary"

    total = len(rows)
    approved = sum(1 for r in rows if r["approved"])
    mismatches = sum(1 for r in rows if r["was_mismatch"])

    ws_sum.append(["WAF Classification History — Summary Report"])
    ws_sum.merge_cells("A1:D1")
    ws_sum.cell(1, 1).font = Font(bold=True, size=16, color="1F4E79")

    ws_sum.append([])
    ws_sum.append(["Generated", datetime.now().strftime("%B %d, %Y %I:%M %p")])
    ws_sum.append(["Total Classifications", total])
    ws_sum.append(["Approved to Ground Truth", approved])
    ws_sum.append(["Mismatches Detected", mismatches])
    ws_sum.append(["Approval Rate", f"{round(approved/total*100,1)}%" if total else "0%"])
    ws_sum.append([])

    # Category summary
    ws_sum.append(["WAF Category", "Count", "% of Total"])
    row_num = ws_sum.max_row
    for col in range(1, 4):
        c = ws_sum.cell(row=row_num, column=col)
        c.font = header_font; c.fill = header_fill; c.alignment = header_align

    cat_counts = defaultdict(int)
    for r in rows:
        cat_counts[r["waf_category"] or "Unknown"] += 1

    for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
        ws_sum.append([cat, cnt, f"{round(cnt/total*100,1)}%" if total else "0%"])

    auto_width(ws_sum, 4)

    # ══════════ Sheet 2: Monthly Rollups ══════════
    ws_month = wb.create_sheet("Monthly Rollups")

    monthly = defaultdict(list)
    for r in rows:
        from datetime import datetime as dt
        ts = dt.fromisoformat(r["timestamp"])
        key = ts.strftime("%Y-%m")
        monthly[key].append(r)

    headers_m = ["Month", "Total", "Approved", "Mismatches", "Approval Rate",
                 "Mismatch Rate", "Run", "Change", "Top Category"]
    ws_month.append(headers_m)
    style_header(ws_month, len(headers_m))

    for month_key in sorted(monthly.keys()):
        mrows = monthly[month_key]
        mt = len(mrows)
        ma = sum(1 for r in mrows if r["approved"])
        mm = sum(1 for r in mrows if r["was_mismatch"])
        rc = defaultdict(int)
        mc = defaultdict(int)
        for r in mrows:
            if r["run_change"]: rc[r["run_change"]] += 1
            mc[r["waf_category"] or "Unknown"] += 1
        top_cat = max(mc, key=mc.get) if mc else "—"

        ws_month.append([
            dt.strptime(month_key, "%Y-%m").strftime("%B %Y"),
            mt, ma, mm,
            f"{round(ma/mt*100,1)}%" if mt else "0%",
            f"{round(mm/mt*100,1)}%" if mt else "0%",
            rc.get("Run", 0), rc.get("Change", 0),
            top_cat
        ])

    auto_width(ws_month, len(headers_m))

    # ══════════ Sheet 3: Raw Data ══════════
    ws_raw = wb.create_sheet("Raw Data")

    headers_r = ["Timestamp", "Story Title", "Description", "WAF Category",
                 "Sub-Category", "WAF Color", "Run/Change", "Confidence",
                 "Mismatch", "Original Tag", "Approved", "Team"]
    ws_raw.append(headers_r)
    style_header(ws_raw, len(headers_r))

    for i, r in enumerate(rows, 2):
        ws_raw.append([
            r["timestamp"], r["story_title"], r["story_description"],
            r["waf_category"], r["waf_subcategory"], r["waf_color"],
            r["run_change"], r["confidence"],
            "Yes" if r["was_mismatch"] else "No",
            r["original_tag"],
            "Yes" if r["approved"] else "No",
            r["team"]
        ])
        # Conditional formatting
        if r["was_mismatch"]:
            for col in range(1, len(headers_r) + 1):
                ws_raw.cell(row=i, column=col).fill = red_fill
        elif r["approved"]:
            for col in range(1, len(headers_r) + 1):
                ws_raw.cell(row=i, column=col).fill = green_fill

    auto_width(ws_raw, len(headers_r))
    ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(headers_r))}{ws_raw.max_row}"

    # ── Write to buffer ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import Response
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=waf-history-report.xlsx"}
    )


def _classify_single_batch(batch, batch_offset, system_prompt, api_key, job_id_short):
    """Classify a single batch of stories. Returns list of result dicts."""
    import re
    client = Anthropic(api_key=api_key)
    batch_prompt = "Classify each story below into the correct WAF category. For EACH story, respond with EXACTLY this format on separate lines:\n\n"
    batch_prompt += "STORY 1: [WAF Category] | [WAF Sub-Category] | [WAF Color] | [Run or Change] | [Confidence: HIGH/MEDIUM/LOW] | [One-line reasoning]\n\n"
    batch_prompt += "Here are the stories:\n\n"
    for j, s in enumerate(batch, 1):
        batch_prompt += f"STORY {j}: {s['title']}"
        if s["description"]:
            batch_prompt += f"\nDescription: {s['description'][:300]}"
        batch_prompt += "\n\n"

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=8000,
            system=system_prompt,
            messages=[{"role": "user", "content": batch_prompt}]
        )
        ai_text = response.content[0].text

        ai_lines = []
        for l in ai_text.split("\n"):
            stripped = l.strip().replace("**", "").replace("*", "")
            if re.match(r'^STORY\s+\d+', stripped, re.IGNORECASE):
                ai_lines.append(stripped)

        results = []
        for j, s in enumerate(batch):
            ai_cat = ai_color = ai_rc = ai_conf = ai_reason = ai_subcat = ""
            if j < len(ai_lines):
                parts = ai_lines[j].split("|")
                first_part = parts[0] if parts else ""
                colon_idx = first_part.find(":")
                if colon_idx >= 0:
                    first_part = first_part[colon_idx + 1:]
                if len(parts) >= 1: ai_cat = first_part.strip()
                if len(parts) >= 2: ai_subcat = parts[1].strip()
                if len(parts) >= 3: ai_color = parts[2].strip()
                if len(parts) >= 4: ai_rc = parts[3].strip()
                if len(parts) >= 5: ai_conf = parts[4].strip().replace("Confidence:", "").strip()
                if len(parts) >= 6: ai_reason = parts[5].strip()

            is_match = (
                s["file_category"].lower().strip() == ai_cat.lower().strip()
                if s["file_category"] and ai_cat else None
            )
            results.append({
                "index": batch_offset + j,
                "title": s["title"], "description": s["description"][:200],
                "team": s["team"], "epic": s["epic"], "parent_feature": s["parent_feature"],
                "timestamp": s["timestamp"], "file_category": s["file_category"],
                "file_color": s["file_color"], "file_run_change": s["file_run_change"],
                "file_subcategory": s.get("file_subcategory", ""), "file_confidence": s.get("file_confidence", ""),
                "ai_category": ai_cat, "ai_subcategory": ai_subcat, "ai_color": ai_color,
                "ai_run_change": ai_rc, "ai_confidence": ai_conf, "ai_reason": ai_reason,
                "is_match": is_match,
            })
        return results
    except Exception as e:
        print(f"[BULK-VERIFY][{job_id_short}] ERROR: {type(e).__name__}: {str(e)[:300]}")
        results = []
        for j, s in enumerate(batch):
            results.append({
                "index": batch_offset + j,
                "title": s["title"], "description": s["description"][:200],
                "team": s["team"], "epic": s["epic"], "parent_feature": s["parent_feature"],
                "timestamp": s["timestamp"], "file_category": s["file_category"],
                "file_color": s["file_color"], "file_run_change": s["file_run_change"],
                "file_subcategory": s.get("file_subcategory", ""), "file_confidence": s.get("file_confidence", ""),
                "ai_category": "", "ai_subcategory": "", "ai_color": "", "ai_run_change": "",
                "ai_confidence": "", "ai_reason": f"API error: {str(e)[:100]}", "is_match": None,
            })
        return results


def _run_verify_job(job_id, stories, filename, ext, row_count):
    """Background thread: classify stories using concurrent workers for speed."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    job = verify_jobs[job_id]

    try:
        system_prompt = build_system_prompt()
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")

        batch_size = 50
        batches = []
        for i in range(0, len(stories), batch_size):
            batches.append((stories[i:i + batch_size], i))

        total_batches = len(batches)
        job["total_batches"] = total_batches

        # Run up to 5 concurrent API calls for speed
        all_results = [None] * total_batches
        completed = [0]  # mutable counter for thread safety
        lock = threading.Lock()

        def process_batch(batch_idx, batch, offset):
            result = _classify_single_batch(batch, offset, system_prompt, api_key, job_id[:8])
            with lock:
                all_results[batch_idx] = result
                completed[0] += 1
                job["completed_batches"] = completed[0]
                job["stories_processed"] = min(completed[0] * batch_size, len(stories))
                print(f"[BULK-VERIFY][{job_id[:8]}] Batch {completed[0]}/{total_batches} done")
            return result

        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {}
            for idx, (batch, offset) in enumerate(batches):
                f = executor.submit(process_batch, idx, batch, offset)
                futures[f] = idx

            for f in as_completed(futures):
                f.result()  # raise any exceptions

        # Flatten results in order
        results = []
        for batch_result in all_results:
            if batch_result:
                results.extend(batch_result)

        matches = sum(1 for r in results if r["is_match"] is True)
        mismatches = sum(1 for r in results if r["is_match"] is False)
        untagged = sum(1 for r in results if r["is_match"] is None)

        # Record in upload history using a fresh DB connection (we're in a thread)
        db = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
        cursor = db.execute(
            """INSERT INTO upload_history (uploaded_at, filename, row_count, imported_count, file_type, status, results_json)
               VALUES (?, ?, ?, ?, ?, 'verified', ?)""",
            (datetime.now().isoformat(), filename, row_count, len(results), ext, json.dumps(results))
        )
        verify_upload_id = cursor.lastrowid
        db.commit()
        db.close()

        job["status"] = "done"
        job["upload_id"] = verify_upload_id
        job["results"] = results
        job["matches"] = matches
        job["mismatches"] = mismatches
        job["untagged"] = untagged
        print(f"[BULK-VERIFY][{job_id[:8]}] COMPLETE: {len(results)} stories, {matches} matches, {mismatches} mismatches, {untagged} untagged")

    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)
        print(f"[BULK-VERIFY][{job_id[:8]}] FATAL ERROR: {str(e)}")


@app.route("/api/bulk-verify", methods=["POST"])
def bulk_verify():
    """Upload a file of stories, start async AI classification, return job_id for polling."""
    client_ip = request.remote_addr or "unknown"
    if not _check_rate_limit(client_ip):
        logger.warning("Rate limit exceeded for bulk-verify from %s", client_ip)
        return jsonify({"error": "Too many requests. Please wait a minute before uploading again."}), 429

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    filename = secure_filename(file.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    try:
        ext = filename.rsplit(".", 1)[-1].lower()
        if ext in ("csv", "tsv"):
            df = pd.read_csv(filepath, sep="\t" if ext == "tsv" else ",")
        elif ext in ("xlsx", "xls"):
            df = pd.read_excel(filepath)
        else:
            return jsonify({"error": "File must be CSV or Excel"}), 400

        df.columns = [c.strip().lower() for c in df.columns]

        def find_col(keywords):
            for col in df.columns:
                if any(kw in col for kw in keywords):
                    return col
            return None

        title_col = find_col(["title", "summary", "story", "name"])
        desc_col = find_col(["desc", "detail", "body", "acceptance"])
        cat_col = find_col(["waf category", "waf_category", "category"])
        color_col = find_col(["waf color", "waf_color", "color"])
        rc_col = find_col(["run/change", "run_change", "run change"])
        subcat_col = find_col(["sub-category", "sub_category", "subcategory", "waf sub"])
        conf_col = find_col(["confidence", "conf"])
        team_col = find_col(["team", "squad", "group"])
        epic_col = find_col(["epic", "initiative", "program"])
        feature_col = find_col(["feature", "parent feature", "parent_feature", "capability"])
        ts_col = find_col(["timestamp", "date", "created", "created_at"])

        if not title_col:
            return jsonify({"error": "File must have a 'Story Title' or 'Summary' column"}), 400

        stories = []
        for _, row in df.iterrows():
            title = str(row.get(title_col, "")).strip()
            if not title or title == "nan":
                continue
            ts = datetime.now().isoformat()
            if ts_col:
                raw_ts = str(row.get(ts_col, "")).strip()
                if raw_ts and raw_ts != "nan":
                    ts = raw_ts
            stories.append({
                "title": title,
                "description": str(row.get(desc_col, "")).strip() if desc_col else "",
                "file_category": str(row.get(cat_col, "")).strip() if cat_col else "",
                "file_color": str(row.get(color_col, "")).strip() if color_col else "",
                "file_run_change": str(row.get(rc_col, "")).strip() if rc_col else "",
                "file_subcategory": str(row.get(subcat_col, "")).strip() if subcat_col else "",
                "file_confidence": str(row.get(conf_col, "")).strip() if conf_col else "",
                "team": str(row.get(team_col, "default")).strip() if team_col else "default",
                "epic": str(row.get(epic_col, "")).strip() if epic_col else "",
                "parent_feature": str(row.get(feature_col, "")).strip() if feature_col else "",
                "timestamp": ts,
            })

        if not stories:
            return jsonify({"error": "No valid stories found in file"}), 400

        # For small files (<=200 stories), process synchronously for speed
        if len(stories) <= 200:
            import re
            client = get_client()
            system_prompt = build_system_prompt()
            results = []
            batch_size = 25
            for i in range(0, len(stories), batch_size):
                batch = stories[i:i + batch_size]
                batch_prompt = "Classify each story below into the correct WAF category. For EACH story, respond with EXACTLY this format on separate lines:\n\n"
                batch_prompt += "STORY 1: [WAF Category] | [WAF Sub-Category] | [WAF Color] | [Run or Change] | [Confidence: HIGH/MEDIUM/LOW] | [One-line reasoning]\n\n"
                batch_prompt += "Here are the stories:\n\n"
                for j, s in enumerate(batch, 1):
                    batch_prompt += f"STORY {j}: {s['title']}"
                    if s["description"]:
                        batch_prompt += f"\nDescription: {s['description'][:300]}"
                    batch_prompt += "\n\n"
                try:
                    print(f"[BULK-VERIFY] Sending batch {i//batch_size + 1} ({len(batch)} stories) to Claude...")
                    response = client.messages.create(
                        model="claude-sonnet-4-5-20250929", max_tokens=6000,
                        system=system_prompt,
                        messages=[{"role": "user", "content": batch_prompt}]
                    )
                    ai_text = response.content[0].text
                    ai_lines = []
                    for l in ai_text.split("\n"):
                        stripped = l.strip().replace("**", "").replace("*", "")
                        if re.match(r'^STORY\s+\d+', stripped, re.IGNORECASE):
                            ai_lines.append(stripped)
                    for j, s in enumerate(batch):
                        ai_cat = ai_color = ai_rc = ai_conf = ai_reason = ai_subcat = ""
                        if j < len(ai_lines):
                            parts = ai_lines[j].split("|")
                            first_part = parts[0] if parts else ""
                            colon_idx = first_part.find(":")
                            if colon_idx >= 0:
                                first_part = first_part[colon_idx + 1:]
                            if len(parts) >= 1: ai_cat = first_part.strip()
                            if len(parts) >= 2: ai_subcat = parts[1].strip()
                            if len(parts) >= 3: ai_color = parts[2].strip()
                            if len(parts) >= 4: ai_rc = parts[3].strip()
                            if len(parts) >= 5: ai_conf = parts[4].strip().replace("Confidence:", "").strip()
                            if len(parts) >= 6: ai_reason = parts[5].strip()
                        is_match = (
                            s["file_category"].lower().strip() == ai_cat.lower().strip()
                            if s["file_category"] and ai_cat else None
                        )
                        results.append({
                            "index": i + j, "title": s["title"], "description": s["description"][:200],
                            "team": s["team"], "epic": s["epic"], "parent_feature": s["parent_feature"],
                            "timestamp": s["timestamp"], "file_category": s["file_category"],
                            "file_color": s["file_color"], "file_run_change": s["file_run_change"],
                            "file_subcategory": s.get("file_subcategory", ""), "file_confidence": s.get("file_confidence", ""),
                            "ai_category": ai_cat, "ai_subcategory": ai_subcat, "ai_color": ai_color,
                            "ai_run_change": ai_rc, "ai_confidence": ai_conf, "ai_reason": ai_reason,
                            "is_match": is_match,
                        })
                except Exception as e:
                    print(f"[BULK-VERIFY] ERROR in batch {i//batch_size + 1}: {type(e).__name__}: {str(e)[:300]}")
                    for s in batch:
                        results.append({
                            "index": len(results), "title": s["title"], "description": s["description"][:200],
                            "team": s["team"], "epic": s["epic"], "parent_feature": s["parent_feature"],
                            "timestamp": s["timestamp"], "file_category": s["file_category"],
                            "file_color": s["file_color"], "file_run_change": s["file_run_change"],
                            "file_subcategory": s.get("file_subcategory", ""), "file_confidence": s.get("file_confidence", ""),
                            "ai_category": "", "ai_subcategory": "", "ai_color": "", "ai_run_change": "",
                            "ai_confidence": "", "ai_reason": f"API error: {str(e)[:100]}", "is_match": None,
                        })
            matches = sum(1 for r in results if r["is_match"] is True)
            mismatches = sum(1 for r in results if r["is_match"] is False)
            untagged = sum(1 for r in results if r["is_match"] is None)
            db = get_db()
            cursor = db.execute(
                """INSERT INTO upload_history (uploaded_at, filename, row_count, imported_count, file_type, status, results_json)
                   VALUES (?, ?, ?, ?, ?, 'verified', ?)""",
                (datetime.now().isoformat(), filename, len(df), len(results), ext, json.dumps(results))
            )
            verify_upload_id = cursor.lastrowid
            db.commit()
            return jsonify({
                "success": True, "filename": filename, "upload_id": verify_upload_id,
                "total": len(results), "matches": matches, "mismatches": mismatches,
                "untagged": untagged, "results": results,
            })

        # For large files (>200 stories), process asynchronously with progress polling
        job_id = str(uuid.uuid4())
        verify_jobs[job_id] = {
            "status": "processing",
            "total_stories": len(stories),
            "stories_processed": 0,
            "total_batches": 0,
            "completed_batches": 0,
            "current_batch": 0,
            "filename": filename,
            "results": None,
            "error": None,
        }

        thread = threading.Thread(
            target=_run_verify_job,
            args=(job_id, stories, filename, ext, len(df)),
            daemon=True
        )
        thread.start()

        return jsonify({
            "success": True,
            "async": True,
            "job_id": job_id,
            "total_stories": len(stories),
            "message": f"Processing {len(stories)} stories in background. Poll /api/bulk-verify/status/{job_id} for progress.",
        })

    except ValueError as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        logger.error("Verification failed: %s", e, exc_info=True)
        return jsonify({"error": "Verification failed. Please try again."}), 500


@app.route("/api/bulk-verify/status/<job_id>", methods=["GET"])
def bulk_verify_status(job_id):
    """Poll progress of an async bulk-verify job."""
    job = verify_jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    resp = {
        "status": job["status"],
        "total_stories": job["total_stories"],
        "stories_processed": job["stories_processed"],
        "total_batches": job["total_batches"],
        "completed_batches": job["completed_batches"],
        "filename": job["filename"],
    }

    if job["status"] == "done":
        resp["success"] = True
        resp["upload_id"] = job.get("upload_id")
        resp["total"] = len(job["results"])
        resp["matches"] = job["matches"]
        resp["mismatches"] = job["mismatches"]
        resp["untagged"] = job["untagged"]
        resp["results"] = job["results"]
        # Clean up after results are fetched
        del verify_jobs[job_id]
    elif job["status"] == "error":
        resp["error"] = job["error"]
        del verify_jobs[job_id]

    return jsonify(resp)


@app.route("/api/bulk-verify/save", methods=["POST"])
def bulk_verify_save():
    """Save approved rows from bulk verification to the database."""
    data = request.json
    if not data or "rows" not in data:
        return jsonify({"error": "No rows provided"}), 400

    saved = 0
    upload_id = data.get("upload_id")
    db = get_db()
    for row in data["rows"]:
        # Determine which category to use (AI recommendation or file's original)
        use_ai = row.get("use_ai", True)
        category = row.get("ai_category", "") if use_ai else row.get("file_category", "")
        subcategory = row.get("ai_subcategory", "") if use_ai else row.get("file_subcategory", "")
        color = row.get("ai_color", "") if use_ai else row.get("file_color", "")
        run_change = row.get("ai_run_change", "") if use_ai else row.get("file_run_change", "")
        confidence = row.get("ai_confidence", "") if use_ai else row.get("file_confidence", "")
        ts = row.get("timestamp", datetime.now().isoformat())

        db.execute(
            """INSERT INTO classifications
               (timestamp, story_title, story_description, waf_category,
                waf_subcategory, waf_color, run_change, confidence,
                was_mismatch, original_tag, approved, team, epic, parent_feature, upload_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?)""",
            (ts,
             row.get("title", ""),
             row.get("description", ""),
             category,
             subcategory,
             color,
             run_change,
             confidence,
             1 if row.get("is_match") is False else 0,
             row.get("file_category", ""),
             row.get("team", "default"),
             row.get("epic", ""),
             row.get("parent_feature", ""),
             upload_id)
        )
        saved += 1

    db.commit()
    return jsonify({"success": True, "saved": saved})


# ── Epic Lineage API ──────────────────────────────────────────────────

@app.route("/lineage")
def lineage_page():
    return send_from_directory("static", "lineage.html")


@app.route("/api/epics", methods=["GET"])
def list_epics():
    """List all unique epics with story counts. Optional ?upload_id= filter."""
    db = get_db()
    upload_id = request.args.get("upload_id", "")
    where = "epic != '' AND epic IS NOT NULL"
    params = []
    if upload_id:
        where += " AND upload_id = ?"
        params.append(int(upload_id))

    rows = db.execute(
        f"""SELECT epic, COUNT(*) as cnt,
                  SUM(CASE WHEN was_mismatch=1 THEN 1 ELSE 0 END) as mismatches,
                  SUM(CASE WHEN approved=1 THEN 1 ELSE 0 END) as approved
           FROM classifications
           WHERE {where}
           GROUP BY epic ORDER BY cnt DESC""",
        params
    ).fetchall()

    return jsonify({
        "epics": [{"name": r["epic"], "story_count": r["cnt"],
                    "mismatches": r["mismatches"], "approved": r["approved"]}
                   for r in rows]
    })


@app.route("/api/epics/summary", methods=["GET"])
def epic_summary():
    """Get dashboard-style summary for all epics or a specific epic. Optional ?upload_id= filter."""
    db = get_db()
    epic_filter = request.args.get("epic", "")
    upload_id = request.args.get("upload_id", "")

    where = "epic != '' AND epic IS NOT NULL"
    params = []
    if epic_filter:
        where += " AND epic = ?"
        params.append(epic_filter)
    if upload_id:
        where += " AND upload_id = ?"
        params.append(int(upload_id))

    rows = db.execute(
        f"""SELECT id, timestamp, story_title, story_description, waf_category,
                   waf_subcategory, waf_color, run_change, confidence,
                   was_mismatch, original_tag, approved, team, epic, parent_feature
            FROM classifications WHERE {where} ORDER BY epic, timestamp DESC""",
        params
    ).fetchall()

    if not rows:
        return jsonify({"epics": []})

    from collections import defaultdict

    epics = defaultdict(list)
    for r in rows:
        epics[r["epic"]].append(r)

    result = []
    for epic_name, stories in epics.items():
        total = len(stories)
        approved = sum(1 for s in stories if s["approved"])
        mismatches = sum(1 for s in stories if s["was_mismatch"])

        cat_counts = {}
        color_counts = {}
        rc_counts = {}
        feature_map = defaultdict(list)  # parent_feature -> stories

        for s in stories:
            cat = s["waf_category"] or "Unknown"
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
            if s["waf_color"]:
                color_counts[s["waf_color"]] = color_counts.get(s["waf_color"], 0) + 1
            if s["run_change"]:
                rc_counts[s["run_change"]] = rc_counts.get(s["run_change"], 0) + 1
            feat = s["parent_feature"] or "(Direct)"
            feature_map[feat].append({
                "id": s["id"],
                "title": s["story_title"],
                "description": s["story_description"],
                "category": s["waf_category"],
                "original_tag": s["original_tag"],
                "color": s["waf_color"],
                "run_change": s["run_change"],
                "confidence": s["confidence"],
                "mismatch": bool(s["was_mismatch"]),
                "approved": bool(s["approved"]),
                "team": s["team"],
            })

        # Build tree: epic -> features -> stories
        features = []
        for feat_name, feat_stories in feature_map.items():
            feat_cats = {}
            for fs in feat_stories:
                c = fs["category"] or "Unknown"
                feat_cats[c] = feat_cats.get(c, 0) + 1
            features.append({
                "name": feat_name,
                "story_count": len(feat_stories),
                "categories": feat_cats,
                "stories": feat_stories,
            })

        # ── Health metrics ──
        # Dominant color: the most frequent WAF color in this epic
        dominant_color = max(color_counts, key=color_counts.get) if color_counts else ""
        dominant_color_pct = round(color_counts.get(dominant_color, 0) / total * 100, 1) if total and dominant_color else 0
        # Color consistency: % of stories matching the dominant color (higher = more focused)
        color_consistency = dominant_color_pct
        # Category focus: % of stories in the top category
        dominant_cat = max(cat_counts, key=cat_counts.get) if cat_counts else ""
        dominant_cat_pct = round(cat_counts.get(dominant_cat, 0) / total * 100, 1) if total and dominant_cat else 0
        # Unique colors & categories
        unique_colors = len(color_counts)
        unique_categories = len(cat_counts)
        # Health score: 0-100 (higher = cleaner)
        # Penalize for: many colors, low dominant %, mismatches
        health = round(
            (dominant_color_pct * 0.4) +
            (dominant_cat_pct * 0.3) +
            ((100 - min(unique_colors * 15, 100)) * 0.2) +
            ((100 - round(mismatches / total * 100, 1) if total else 100) * 0.1)
        ) if total else 0
        health = max(0, min(100, health))

        # Flag: is this epic "mixed" (3+ colors or dominant < 60%)?
        is_mixed = unique_colors >= 3 or dominant_color_pct < 60

        # Teams involved
        team_set = set(s["team"] for s in stories if s["team"] and s["team"] != "default")

        result.append({
            "epic": epic_name,
            "total_stories": total,
            "approved": approved,
            "mismatches": mismatches,
            "approval_rate": round(approved / total * 100, 1) if total else 0,
            "mismatch_rate": round(mismatches / total * 100, 1) if total else 0,
            "categories": cat_counts,
            "colors": color_counts,
            "run_change": rc_counts,
            "features": features,
            "health": health,
            "dominant_color": dominant_color,
            "dominant_color_pct": dominant_color_pct,
            "dominant_category": dominant_cat,
            "dominant_category_pct": dominant_cat_pct,
            "color_consistency": color_consistency,
            "unique_colors": unique_colors,
            "unique_categories": unique_categories,
            "is_mixed": is_mixed,
            "teams": list(team_set),
        })

    return jsonify({"epics": result})


@app.route("/api/epics/uploads", methods=["GET"])
def epic_uploads():
    """List uploads that contain epic data, for the lineage filter dropdown."""
    db = get_db()
    rows = db.execute(
        """SELECT DISTINCT c.upload_id, h.filename, h.uploaded_at, h.imported_count
           FROM classifications c
           JOIN upload_history h ON h.id = c.upload_id
           WHERE c.epic != '' AND c.epic IS NOT NULL AND c.upload_id IS NOT NULL
           GROUP BY c.upload_id
           ORDER BY h.uploaded_at DESC"""
    ).fetchall()
    return jsonify({
        "uploads": [{"upload_id": r["upload_id"], "filename": r["filename"],
                      "uploaded_at": r["uploaded_at"], "imported_count": r["imported_count"]}
                     for r in rows]
    })


@app.route("/api/epics/assign", methods=["POST"])
def assign_epic():
    """Assign or update epic/parent_feature for one or more classifications."""
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    ids = data.get("ids", [])
    epic = data.get("epic", "").strip()
    parent_feature = data.get("parent_feature", "").strip()

    if not ids or not epic:
        return jsonify({"error": "ids and epic are required"}), 400

    db = get_db()
    placeholders = ",".join("?" for _ in ids)
    db.execute(
        f"UPDATE classifications SET epic=?, parent_feature=? WHERE id IN ({placeholders})",
        [epic, parent_feature] + ids
    )
    db.commit()

    return jsonify({"success": True, "updated": len(ids)})


@app.route("/api/epics/autocomplete", methods=["GET"])
def epic_autocomplete():
    """Get autocomplete suggestions for epic names."""
    db = get_db()
    q = request.args.get("q", "").strip()

    if q:
        rows = db.execute(
            "SELECT DISTINCT epic FROM classifications WHERE epic LIKE ? AND epic != '' ORDER BY epic LIMIT 20",
            (f"%{q}%",)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT DISTINCT epic FROM classifications WHERE epic != '' ORDER BY epic LIMIT 50"
        ).fetchall()

    # Also get distinct parent features
    if q:
        feat_rows = db.execute(
            "SELECT DISTINCT parent_feature FROM classifications WHERE parent_feature LIKE ? AND parent_feature != '' ORDER BY parent_feature LIMIT 20",
            (f"%{q}%",)
        ).fetchall()
    else:
        feat_rows = db.execute(
            "SELECT DISTINCT parent_feature FROM classifications WHERE parent_feature != '' ORDER BY parent_feature LIMIT 50"
        ).fetchall()

    return jsonify({
        "epics": [r["epic"] for r in rows],
        "features": [r["parent_feature"] for r in feat_rows],
    })


@app.route("/api/dashboard/save", methods=["POST"])
def dashboard_save():
    """Manually save a classification from the chat to history."""
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    try:
        row_id = save_classification(
            title=data.get("title", ""),
            description=data.get("description", ""),
            category=data.get("waf_category", ""),
            subcategory=data.get("waf_subcategory", ""),
            color=data.get("waf_color", ""),
            run_change=data.get("run_change", ""),
            confidence=data.get("confidence", ""),
            was_mismatch=data.get("was_mismatch", False),
            original_tag=data.get("original_tag", ""),
            approved=data.get("approved", False),
        )
        return jsonify({"success": True, "id": row_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def auto_load_sample_data():
    """Auto-load WAF definitions and ground truth from sample-data/ on startup."""
    sample_dir = os.path.join(os.path.dirname(__file__), "sample-data")

    # Load WAF definitions
    waf_file = os.path.join(sample_dir, "waf-definitions.csv")
    if os.path.exists(waf_file):
        try:
            text, categories, df = parse_waf_file(waf_file, "waf-definitions.csv")
            waf_store["definitions"] = df  # Store DataFrame for /api/waf-definitions
            waf_store["raw_text"] = text
            waf_store["filename"] = "waf-definitions.csv"
            waf_store["categories"] = categories
            print(f"  Auto-loaded WAF definitions: {len(categories)} categories")
        except Exception as e:
            print(f"  Warning: Failed to auto-load WAF definitions: {e}")

    # Load ground truth
    gt_file = os.path.join(sample_dir, "sample-ground-truth.csv")
    if os.path.exists(gt_file):
        try:
            examples, stats, col_map = parse_ground_truth(gt_file, "sample-ground-truth.csv")
            ground_truth_store["loaded"] = True
            ground_truth_store["filename"] = "sample-ground-truth.csv"
            ground_truth_store["examples"] = examples
            ground_truth_store["example_count"] = len(examples)
            ground_truth_store["stats"] = stats
            ground_truth_store["raw_text"] = f"{len(examples)} examples across {len(stats)} categories"
            print(f"  Auto-loaded ground truth: {len(examples)} examples across {len(stats)} categories")
        except Exception as e:
            print(f"  Warning: Failed to auto-load ground truth: {e}")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    print(f"\n{'='*60}")
    print(f"  WAF Category Classifier")
    print(f"  Running at: http://localhost:{port}")
    print(f"  Analytics:  http://localhost:{port}/history")
    print(f"  API Key configured: {bool(os.environ.get('ANTHROPIC_API_KEY'))}")
    init_db()
    print(f"  Database initialized: {DB_PATH}")
    auto_load_sample_data()
    print(f"{'='*60}\n")
    app.run(host="0.0.0.0", port=port, debug=False)
