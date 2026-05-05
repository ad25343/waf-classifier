"""
WAF Classifier — Analytics, Dashboard, and History API routes.
Blueprint: analytics_bp
"""

import os
import csv
import io
import json
from datetime import datetime
from collections import defaultdict

import pandas as pd
from flask import Blueprint, request, jsonify, Response, g
from werkzeug.utils import secure_filename

from database import get_db, save_classification
from waf_core import normalize_waf_category
from state import waf_store
from config import UPLOAD_FOLDER

analytics_bp = Blueprint("analytics_bp", __name__)


@analytics_bp.route("/api/dashboard/summary", methods=["GET"])
def dashboard_summary():
    """Get summary stats for the dashboard. Optional ?upload_id= filter."""
    db = get_db()
    uid = request.args.get("upload_id", "")
    wh = " WHERE upload_id = ?" if uid else ""
    wh_and = " AND upload_id = ?" if uid else ""
    params = [int(uid)] if uid else []

    total = db.execute(f"SELECT COUNT(*) FROM classifications{wh}", params).fetchone()[0]
    approved = db.execute(f"SELECT COUNT(*) FROM classifications WHERE approved=1{wh_and}", params).fetchone()[0]
    mismatches = db.execute(f"SELECT COUNT(*) FROM classifications WHERE was_mismatch=1{wh_and}", params).fetchone()[0]

    categories = db.execute(
        f"SELECT waf_category, COUNT(*) as cnt FROM classifications{wh} GROUP BY waf_category ORDER BY cnt DESC", params
    ).fetchall()

    colors = db.execute(
        f"SELECT waf_color, COUNT(*) as cnt FROM classifications WHERE waf_color != ''{wh_and} GROUP BY waf_color ORDER BY cnt DESC", params
    ).fetchall()

    confidence = db.execute(
        f"SELECT confidence, COUNT(*) as cnt FROM classifications WHERE confidence != ''{wh_and} GROUP BY confidence ORDER BY cnt DESC", params
    ).fetchall()

    submitted_waf_rows = db.execute(
        f"SELECT original_tag, COUNT(*) as cnt FROM classifications WHERE original_tag != ''{wh_and} GROUP BY original_tag ORDER BY cnt DESC", params
    ).fetchall()
    # Most common original_color per original_tag
    orig_color_rows = db.execute(
        f"""SELECT original_tag, original_color, COUNT(*) as cnt
            FROM classifications
            WHERE original_tag != '' AND original_color != ''{wh_and}
            GROUP BY original_tag, original_color""", params
    ).fetchall()
    orig_color_map = {}
    for r in orig_color_rows:
        tag = r["original_tag"]
        if tag not in orig_color_map or r["cnt"] > orig_color_map[tag][1]:
            orig_color_map[tag] = (r["original_color"], r["cnt"])

    run_change = db.execute(
        f"SELECT run_change, COUNT(*) as cnt FROM classifications WHERE run_change != ''{wh_and} GROUP BY run_change ORDER BY cnt DESC", params
    ).fetchall()

    daily = db.execute(
        f"""SELECT DATE(timestamp) as day, COUNT(*) as cnt
           FROM classifications{wh}
           GROUP BY DATE(timestamp)
           ORDER BY day DESC LIMIT 30""", params
    ).fetchall()

    recent = db.execute(
        f"""SELECT id, timestamp, story_title, waf_category, waf_color,
                  confidence, was_mismatch, approved
           FROM classifications{wh} ORDER BY id DESC LIMIT 20""", params
    ).fetchall()

    return jsonify({
        "total_classifications": total,
        "total_approved": approved,
        "total_mismatches": mismatches,
        "approval_rate": round(approved / total * 100, 1) if total > 0 else 0,
        "categories": [{"name": r["waf_category"], "count": r["cnt"]} for r in categories],
        "colors": [{"name": r["waf_color"], "count": r["cnt"]} for r in colors],
        "confidence": [{"name": r["confidence"], "count": r["cnt"]} for r in confidence],
        "submitted_waf": [{"name": r["original_tag"], "count": r["cnt"], "original_color": orig_color_map.get(r["original_tag"], ("", 0))[0]} for r in submitted_waf_rows],
        "run_change": [{"name": r["run_change"], "count": r["cnt"]} for r in run_change],
        "daily_trend": [{"date": r["day"], "count": r["cnt"]} for r in reversed(list(daily))],
        "recent": [{
            "id": r["id"], "timestamp": r["timestamp"], "title": r["story_title"],
            "category": r["waf_category"], "color": r["waf_color"],
            "confidence": r["confidence"], "mismatch": bool(r["was_mismatch"]),
            "approved": bool(r["approved"])
        } for r in recent]
    })


@analytics_bp.route("/api/dashboard/stories", methods=["GET"])
def dashboard_stories():
    """Get filtered story list for drill-down.
    Supports ?filter=mismatches|category|color|confidence&value=X&upload_id=Y&q=text
    q= uses FTS5 when present; falls back to LIKE on story_title.
    """
    db = get_db()
    uid   = request.args.get("upload_id", "")
    filt  = request.args.get("filter", "")
    value = request.args.get("value", "")
    q     = request.args.get("q", "").strip()
    page     = max(1, int(request.args.get("page", "1")))
    per_page_raw = request.args.get("per_page", "100")
    per_page = 9999 if per_page_raw == "all" else min(5000, max(1, int(per_page_raw)))

    where_clauses = []
    params = []

    if uid:
        where_clauses.append("upload_id = ?")
        params.append(int(uid))

    if filt == "mismatches":
        where_clauses.append("was_mismatch = 1")
    elif filt == "approved":
        where_clauses.append("approved = 1")
    elif filt == "category" and value:
        where_clauses.append("waf_category = ?")
        params.append(value)
    elif filt == "color" and value:
        where_clauses.append("waf_color = ?")
        params.append(value)
    elif filt == "confidence" and value:
        where_clauses.append("confidence = ?")
        params.append(value)
    elif filt == "run_change" and value:
        where_clauses.append("run_change = ?")
        params.append(value)

    # Column-level keyword filters (q_title, q_epic, q_waf, q_ai, q_color, q_status)
    col_filters = {
        "q_title":  ("story_title",   False),
        "q_epic":   ("epic",          False),
        "q_waf":    ("original_tag",  False),
        "q_ai":     ("waf_category",  False),
        "q_color":  ("waf_color",     False),
        "q_team":   ("team",          False),
    }
    for param_name, (col_name, _) in col_filters.items():
        v = request.args.get(param_name, "").strip()
        if v:
            where_clauses.append(f"LOWER({col_name}) LIKE ?")
            params.append(f"%{v.lower()}%")

    q_status = request.args.get("q_status", "").strip().lower()
    if q_status:
        if "mismatch" in q_status:
            where_clauses.append("was_mismatch = 1")
        elif "match" in q_status or "correct" in q_status:
            where_clauses.append("was_mismatch = 0")

    # Full-text search via FTS5 (restricts to matching rowids)
    fts_ids = None
    if q:
        try:
            fts_term = q.replace('"', '""')
            fts_rows = db.execute(
                "SELECT rowid FROM classifications_fts WHERE classifications_fts MATCH ?",
                [f'"{fts_term}"']
            ).fetchall()
            fts_ids = [r[0] for r in fts_rows]
            if not fts_ids:
                # Fallback: prefix match
                fts_rows = db.execute(
                    "SELECT rowid FROM classifications_fts WHERE classifications_fts MATCH ?",
                    [fts_term + "*"]
                ).fetchall()
                fts_ids = [r[0] for r in fts_rows]
        except Exception:
            fts_ids = None

    if fts_ids is not None:
        if fts_ids:
            placeholders = ",".join("?" * len(fts_ids))
            where_clauses.append(f"id IN ({placeholders})")
            params.extend(fts_ids)
        else:
            # No FTS matches — return empty result
            return jsonify({"total": 0, "page": page, "per_page": per_page,
                            "total_pages": 0, "stories": []})

    where = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""

    # Safe server-side sort (whitelist only)
    SORT_COLS = {
        "timestamp": "timestamp", "title": "story_title",
        "category": "waf_category", "color": "waf_color",
        "confidence": "confidence", "status": "was_mismatch",
        "epic": "epic", "original_tag": "original_tag",
        "run_change": "run_change", "team": "team",
    }
    sort_col = SORT_COLS.get(request.args.get("sort", "timestamp"), "timestamp")
    sort_dir = "ASC" if request.args.get("dir", "desc").lower() == "asc" else "DESC"

    total = db.execute(f"SELECT COUNT(*) FROM classifications{where}", params).fetchone()[0]

    offset = (page - 1) * per_page
    rows = db.execute(
        f"""SELECT id, story_title, story_description, waf_category, original_tag,
                   waf_color, run_change, confidence, was_mismatch, approved, team,
                   epic, parent_feature, timestamp, original_color,
                   story_id, story_points, waf_reasoning, pi_number
            FROM classifications{where}
            ORDER BY {sort_col} {sort_dir} LIMIT ? OFFSET ?""",
        params + [per_page, offset]
    ).fetchall()

    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page,
        "stories": [{
            "id": r["id"], "title": r["story_title"], "description": (r["story_description"] or "")[:200],
            "category": r["waf_category"], "original_tag": r["original_tag"],
            "color": r["waf_color"], "run_change": r["run_change"],
            "confidence": r["confidence"], "mismatch": bool(r["was_mismatch"]),
            "approved": bool(r["approved"]), "team": r["team"],
            "epic": r["epic"], "feature": r["parent_feature"],
            "timestamp": r["timestamp"],
            "original_color": r["original_color"] or "",
            "story_id": r["story_id"] or "",
            "story_points": r["story_points"] or "",
            "waf_reasoning": r["waf_reasoning"] or "",
            "pi_number": r["pi_number"] or "",
        } for r in rows]
    })


@analytics_bp.route("/api/dashboard/save", methods=["POST"])
def dashboard_save():
    """Manually save a classification from the chat to history."""
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400

    try:
        row_id = save_classification(
            title=data.get("title", ""),
            description=data.get("description", ""),
            category=data.get("waf_category", ""),
            team_of_teams=data.get("team_of_teams", ""),
            color=data.get("waf_color", ""),
            run_change=data.get("run_change", ""),
            confidence=data.get("confidence", ""),
            was_mismatch=data.get("was_mismatch", False),
            original_tag=data.get("original_tag", ""),
            approved=data.get("approved", False),
            pi_number=data.get("pi_number", ""),
        )
        return jsonify({"success": True, "id": row_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@analytics_bp.route("/api/history/sprints", methods=["GET"])
def history_sprints():
    """Get sprint-over-sprint classification trends.
    Sprints are 2-week windows. Query params: ?sprints=10&upload_id="""
    db = get_db()
    num_sprints = int(request.args.get("sprints", 10))
    uid = request.args.get("upload_id", "")
    wh = " WHERE upload_id = ?" if uid else ""
    params = [int(uid)] if uid else []

    rows = db.execute(
        f"""SELECT timestamp, waf_category, waf_color, run_change, confidence,
                  was_mismatch, approved, team
           FROM classifications{wh} ORDER BY timestamp ASC""",
        params
    ).fetchall()

    if not rows:
        return jsonify({"sprints": []})

    from datetime import datetime as dt, timedelta

    # Determine sprint boundaries (2-week windows from the earliest record)
    first_ts = dt.fromisoformat(rows[0]["timestamp"])
    # Align to Monday of that week
    sprint_start = first_ts - timedelta(days=first_ts.weekday())
    sprint_start = sprint_start.replace(hour=0, minute=0, second=0, microsecond=0)

    sprints = []
    now = dt.now()
    while sprint_start <= now:
        sprint_end = sprint_start + timedelta(days=14)
        sprint_rows = [r for r in rows
                       if sprint_start <= dt.fromisoformat(r["timestamp"]) < sprint_end]

        if sprint_rows:
            total = len(sprint_rows)
            approved = sum(1 for r in sprint_rows if r["approved"])
            mismatches = sum(1 for r in sprint_rows if r["was_mismatch"])

            # Category breakdown
            cat_counts = {}
            color_counts = {}
            rc_counts = {}
            conf_counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
            for r in sprint_rows:
                cat = r["waf_category"] or "Unknown"
                cat_counts[cat] = cat_counts.get(cat, 0) + 1
                if r["waf_color"]:
                    color_counts[r["waf_color"]] = color_counts.get(r["waf_color"], 0) + 1
                if r["run_change"]:
                    rc_counts[r["run_change"]] = rc_counts.get(r["run_change"], 0) + 1
                c = (r["confidence"] or "").upper()
                if c in conf_counts:
                    conf_counts[c] += 1

            sprints.append({
                "sprint_label": f"{sprint_start.strftime('%b %d')} \u2013 {(sprint_end - timedelta(days=1)).strftime('%b %d, %Y')}",
                "start": sprint_start.isoformat(),
                "end": sprint_end.isoformat(),
                "total": total,
                "approved": approved,
                "mismatches": mismatches,
                "approval_rate": round(approved / total * 100, 1) if total else 0,
                "mismatch_rate": round(mismatches / total * 100, 1) if total else 0,
                "categories": cat_counts,
                "colors": color_counts,
                "run_change": rc_counts,
                "confidence": conf_counts,
            })

        sprint_start = sprint_end

    # Return only the last N sprints
    return jsonify({"sprints": sprints[-num_sprints:]})


@analytics_bp.route("/api/history/monthly", methods=["GET"])
def history_monthly():
    """Get monthly rollup reports with period-over-period comparisons. Optional ?upload_id="""
    db = get_db()
    num_months = int(request.args.get("months", 12))
    uid = request.args.get("upload_id", "")
    wh = " WHERE upload_id = ?" if uid else ""
    params = [int(uid)] if uid else []

    rows = db.execute(
        f"""SELECT timestamp, waf_category, waf_color, run_change, confidence,
                  was_mismatch, approved, team
           FROM classifications{wh} ORDER BY timestamp ASC""",
        params
    ).fetchall()

    if not rows:
        return jsonify({"months": []})

    from datetime import datetime as dt

    monthly = defaultdict(list)
    for r in rows:
        ts = dt.fromisoformat(r["timestamp"])
        key = ts.strftime("%Y-%m")
        monthly[key].append(r)

    result = []
    sorted_months = sorted(monthly.keys())[-num_months:]
    prev = None

    for month_key in sorted_months:
        month_rows = monthly[month_key]
        total = len(month_rows)
        approved = sum(1 for r in month_rows if r["approved"])
        mismatches = sum(1 for r in month_rows if r["was_mismatch"])

        cat_counts = {}
        color_counts = {}
        rc_counts = {}
        for r in month_rows:
            cat = r["waf_category"] or "Unknown"
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
            if r["waf_color"]:
                color_counts[r["waf_color"]] = color_counts.get(r["waf_color"], 0) + 1
            if r["run_change"]:
                rc_counts[r["run_change"]] = rc_counts.get(r["run_change"], 0) + 1

        # Period-over-period deltas
        delta_total = total - prev["total"] if prev else 0
        delta_mismatches = (round(mismatches / total * 100, 1) if total else 0) - (prev["mismatch_rate"] if prev else 0)

        entry = {
            "month": month_key,
            "month_label": dt.strptime(month_key, "%Y-%m").strftime("%B %Y"),
            "total": total,
            "approved": approved,
            "mismatches": mismatches,
            "approval_rate": round(approved / total * 100, 1) if total else 0,
            "mismatch_rate": round(mismatches / total * 100, 1) if total else 0,
            "categories": cat_counts,
            "colors": color_counts,
            "run_change": rc_counts,
            "delta_total": delta_total,
            "delta_mismatch_rate": round(delta_mismatches, 1),
        }
        result.append(entry)
        prev = entry

    return jsonify({"months": result})


@analytics_bp.route("/api/history/timeline", methods=["GET"])
def history_timeline():
    """Get full timeline with optional filters.
    Query params: ?from=2025-01-01&to=2025-12-31&team=&category=&color=&page=1&per_page=50"""
    db = get_db()

    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    team = request.args.get("team", "")
    category = request.args.get("category", "")
    color = request.args.get("color", "")
    confidence = request.args.get("confidence", "")
    mismatch_only = request.args.get("mismatch_only", "")
    page = max(1, int(request.args.get("page", 1)))
    per_page = min(500, max(1, int(request.args.get("per_page", 50))))

    uid = request.args.get("upload_id", "")

    where = []
    params = []

    if uid:
        where.append("upload_id = ?")
        params.append(int(uid))
    if date_from:
        where.append("timestamp >= ?")
        params.append(date_from)
    if date_to:
        where.append("timestamp <= ?")
        params.append(date_to + "T23:59:59")
    if team:
        where.append("team = ?")
        params.append(team)
    if category:
        where.append("waf_category = ?")
        params.append(category)
    if color:
        where.append("waf_color = ?")
        params.append(color)
    if confidence:
        where.append("confidence = ?")
        params.append(confidence)
    if mismatch_only == "1":
        where.append("was_mismatch = 1")

    where_clause = " AND ".join(where) if where else "1=1"

    total = db.execute(
        f"SELECT COUNT(*) FROM classifications WHERE {where_clause}", params
    ).fetchone()[0]

    rows = db.execute(
        f"""SELECT id, timestamp, story_title, story_description, waf_category,
                   team_of_teams, waf_color, run_change, confidence,
                   was_mismatch, original_tag, approved, team
            FROM classifications
            WHERE {where_clause}
            ORDER BY timestamp DESC
            LIMIT ? OFFSET ?""",
        params + [per_page, (page - 1) * per_page]
    ).fetchall()

    # Get filter options
    all_categories = db.execute(
        "SELECT DISTINCT waf_category FROM classifications WHERE waf_category != '' ORDER BY waf_category"
    ).fetchall()
    all_colors = db.execute(
        "SELECT DISTINCT waf_color FROM classifications WHERE waf_color != '' ORDER BY waf_color"
    ).fetchall()
    all_teams = db.execute(
        "SELECT DISTINCT team FROM classifications WHERE team != '' ORDER BY team"
    ).fetchall()

    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page if per_page else 1,
        "items": [{
            "id": r["id"],
            "timestamp": r["timestamp"],
            "title": r["story_title"],
            "description": r["story_description"],
            "category": r["waf_category"],
            "team_of_teams": r["team_of_teams"],
            "color": r["waf_color"],
            "run_change": r["run_change"],
            "confidence": r["confidence"],
            "mismatch": bool(r["was_mismatch"]),
            "original_tag": r["original_tag"],
            "approved": bool(r["approved"]),
            "team": r["team"],
        } for r in rows],
        "filters": {
            "categories": [r["waf_category"] for r in all_categories],
            "colors": [r["waf_color"] for r in all_colors],
            "teams": [r["team"] for r in all_teams],
        }
    })


@analytics_bp.route("/api/history/import", methods=["POST"])
def history_import():
    """Bulk import classifications from a CSV/Excel file.
    This lets PMOs import data for teams that don't use the classifier directly.
    Expected columns: Story Title, WAF Category, WAF Color, Run/Change, Confidence, Team."""
    import logging
    logger = logging.getLogger(__name__)

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    filename = secure_filename(file.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    try:
        ext = filename.rsplit(".", 1)[-1].lower()
        if ext in ("csv", "tsv"):
            df = pd.read_csv(filepath, sep="\t" if ext == "tsv" else ",")
        elif ext in ("xlsx", "xls"):
            df = pd.read_excel(filepath)
        else:
            return jsonify({"error": "File must be CSV or Excel"}), 400

        df.columns = [c.strip().lower() for c in df.columns]

        # Match keyword-first (priority order), not column-first.
        # Otherwise a fallback keyword like "subcategory" can win over the
        # primary "team of teams" if the subcategory column appears earlier
        # in the file — which causes the Team of Teams filter to display the
        # wrong data on the Teams view.
        def find_col(kws):
            for kw in kws:
                for col in df.columns:
                    if kw in col:
                        return col
            return None

        title_col        = find_col(["story title", "story name", "title", "summary", "story", "name"])
        desc_col         = find_col(["story description", "description", "desc", "detail", "body"])
        cat_col          = find_col(["waf category", "waf_category", "category"])
        color_col        = find_col(["waf color", "waf_color", "color"])
        rc_col           = find_col(["run/change", "run_change", "run change"])
        tot_col          = find_col(["team of teams", "team_of_teams", "sub-category", "sub_category", "subcategory", "waf sub"])
        conf_col         = find_col(["confidence", "conf"])
        team_col         = find_col(["assigned teams", "assigned team", "assigned_team", "team", "squad", "group"])
        epic_col         = find_col(["epic name", "epic", "initiative", "program"])
        feature_col      = find_col(["feature name", "feature", "parent feature", "parent_feature", "capability"])
        ts_col           = find_col(["timestamp", "time stamp", "date", "created", "created_at"])
        story_id_col     = find_col(["story id", "story_id", "issue key", "issue_key", "ticket", "jira id"])
        story_points_col = find_col(["story points", "story_points", "points", "estimate"])
        pi_number_col    = find_col(["pi name", "pi number", "pi_number", "pi #", "program increment", " pi "])
        epic_id_col      = find_col(["epic id", "epic_id", "epic key", "epic_key", "epic link", "initiative id"])
        feature_id_col   = find_col(["feature id", "feature_id", "feature key", "parent id", "parent_id"])

        if not title_col or not cat_col:
            return jsonify({"error": "File must have at least 'Story Title' (or 'Story Name') and 'WAF Category' columns"}), 400

        db = get_db()

        # Create upload_history record FIRST so we get an upload_id
        cursor = db.execute(
            """INSERT INTO upload_history (uploaded_at, filename, row_count, imported_count, file_type, status)
               VALUES (?, ?, ?, 0, ?, 'importing')""",
            (datetime.now().isoformat(), filename, len(df), ext)
        )
        upload_id = cursor.lastrowid
        db.commit()

        imported = 0
        for _, row in df.iterrows():
            title = str(row.get(title_col, "")).strip()
            if not title or title == "nan":
                continue

            # Use CSV timestamp if available, otherwise now()
            ts = datetime.now().isoformat()
            if ts_col:
                raw_ts = str(row.get(ts_col, "")).strip()
                if raw_ts and raw_ts != "nan":
                    ts = raw_ts

            def _v(col, default=""):
                return str(row.get(col, default)).strip() if col else default

            db.execute(
                """INSERT INTO classifications
                   (timestamp, story_title, story_description, waf_category,
                    team_of_teams, waf_color, run_change, confidence,
                    was_mismatch, original_tag, approved, team, epic, parent_feature,
                    story_id, story_points, pi_number, epic_id, feature_id, upload_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, '', 1, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (ts, title,
                 _v(desc_col), _v(cat_col), _v(tot_col), _v(color_col),
                 _v(rc_col), _v(conf_col), _v(team_col, "default"),
                 _v(epic_col), _v(feature_col),
                 _v(story_id_col), _v(story_points_col),
                 _v(pi_number_col), _v(epic_id_col), _v(feature_id_col),
                 upload_id)
            )
            imported += 1
        db.commit()

        # Update upload_history with final count and status
        db.execute(
            "UPDATE upload_history SET imported_count = ?, status = 'completed' WHERE id = ?",
            (imported, upload_id)
        )
        db.commit()

        return jsonify({"success": True, "imported": imported, "filename": filename, "upload_id": upload_id})
    except Exception as e:
        logger.error("Import failed: %s", e, exc_info=True)
        return jsonify({"error": "Import failed. Check the file format and try again."}), 500


@analytics_bp.route("/api/history/uploads", methods=["GET"])
def get_upload_history():
    """Get list of past file uploads for analytics."""
    db = get_db()
    rows = db.execute(
        """SELECT u.id, u.uploaded_at, u.filename, u.row_count, u.imported_count,
                  u.file_type, u.status,
                  (SELECT COUNT(*) FROM classifications WHERE upload_id = u.id) AS saved_count,
                  CASE WHEN u.results_json IS NOT NULL THEN 1 ELSE 0 END AS has_results
           FROM upload_history u ORDER BY u.uploaded_at DESC LIMIT 20"""
    ).fetchall()
    uploads = [
        {
            "id": r["id"],
            "uploaded_at": r["uploaded_at"],
            "filename": r["filename"],
            "row_count": r["row_count"],
            "imported_count": r["imported_count"],
            "file_type": r["file_type"],
            "status": r["status"],
            "saved_count": r["saved_count"],
            "has_results": bool(r["has_results"])
        }
        for r in rows
    ]
    return jsonify({"uploads": uploads})


@analytics_bp.route("/api/history/uploads/<int:upload_id>", methods=["DELETE"])
def delete_upload(upload_id):
    """Delete an upload and all its associated classifications."""
    import logging
    logger = logging.getLogger(__name__)

    db = get_db()
    row = db.execute("SELECT id, filename FROM upload_history WHERE id = ?", (upload_id,)).fetchone()
    if not row:
        return jsonify({"error": "Upload not found"}), 404

    deleted_count = db.execute("DELETE FROM classifications WHERE upload_id = ?", (upload_id,)).rowcount
    db.execute("DELETE FROM upload_history WHERE id = ?", (upload_id,))
    db.commit()
    logger.info("Deleted upload %d (%s) and %d classifications", upload_id, row["filename"], deleted_count)
    return jsonify({"success": True, "deleted_classifications": deleted_count})


@analytics_bp.route("/api/history/uploads/<int:upload_id>/reload", methods=["POST"])
def reload_upload(upload_id):
    """Reload a previously uploaded file's AI results for re-review and saving."""
    db = get_db()
    row = db.execute(
        "SELECT filename, row_count, results_json FROM upload_history WHERE id = ?",
        (upload_id,)
    ).fetchone()
    if not row:
        return jsonify({"error": "Upload not found"}), 404

    if row["results_json"]:
        results = json.loads(row["results_json"])
        matches   = sum(1 for r in results if r.get("is_match") is True)
        mismatches = sum(1 for r in results if r.get("is_match") is False)
        untagged  = sum(1 for r in results if r.get("is_match") is None)
        return jsonify({
            "success": True, "has_results": True,
            "filename": row["filename"], "upload_id": upload_id,
            "total": len(results), "matches": matches,
            "mismatches": mismatches, "untagged": untagged,
            "results": results,
        })

    return jsonify({"success": True, "has_results": False,
                    "filename": row["filename"],
                    "message": "No saved AI results for this upload."})


@analytics_bp.route("/api/history/export", methods=["GET"])
def history_export():
    """Export filtered history as CSV. Same filters as timeline endpoint."""
    db = get_db()

    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    team = request.args.get("team", "")
    category = request.args.get("category", "")
    color = request.args.get("color", "")

    where = []
    params = []
    if date_from:
        where.append("timestamp >= ?")
        params.append(date_from)
    if date_to:
        where.append("timestamp <= ?")
        params.append(date_to + "T23:59:59")
    if team:
        where.append("team = ?")
        params.append(team)
    if category:
        where.append("waf_category = ?")
        params.append(category)
    if color:
        where.append("waf_color = ?")
        params.append(color)

    where_clause = " AND ".join(where) if where else "1=1"

    rows = db.execute(
        f"""SELECT timestamp, story_title, story_description, waf_category,
                   team_of_teams, waf_color, run_change, confidence,
                   was_mismatch, original_tag, approved, team
            FROM classifications WHERE {where_clause} ORDER BY timestamp DESC""",
        params
    ).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Timestamp", "Story Title", "Description", "WAF Category",
                     "Team of Teams", "WAF Color", "Run/Change", "Confidence",
                     "Mismatch", "Original Tag", "Approved", "Team"])
    for r in rows:
        writer.writerow([r["timestamp"], r["story_title"], r["story_description"],
                         r["waf_category"], r["team_of_teams"], r["waf_color"],
                         r["run_change"], r["confidence"],
                         "Yes" if r["was_mismatch"] else "No",
                         r["original_tag"], "Yes" if r["approved"] else "No",
                         r["team"]])

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=waf-history-export.csv"}
    )


@analytics_bp.route("/api/history/export-xlsx", methods=["GET"])
def history_export_xlsx():
    """Export formatted Excel workbook with Summary, Sprint Trends, Monthly, and Raw Data sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db()

    # Apply same filters as timeline
    date_from = request.args.get("from", "")
    date_to = request.args.get("to", "")
    team = request.args.get("team", "")
    category = request.args.get("category", "")
    color = request.args.get("color", "")

    where = []
    params = []
    if date_from:
        where.append("timestamp >= ?"); params.append(date_from)
    if date_to:
        where.append("timestamp <= ?"); params.append(date_to + "T23:59:59")
    if team:
        where.append("team = ?"); params.append(team)
    if category:
        where.append("waf_category = ?"); params.append(category)
    if color:
        where.append("waf_color = ?"); params.append(color)

    where_clause = " AND ".join(where) if where else "1=1"

    rows = db.execute(
        f"""SELECT timestamp, story_title, story_description, waf_category,
                   team_of_teams, waf_color, run_change, confidence,
                   was_mismatch, original_tag, approved, team
            FROM classifications WHERE {where_clause} ORDER BY timestamp DESC""",
        params
    ).fetchall()

    wb = Workbook()

    # -- Styles --
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    def style_header(ws, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_width(ws, col_count, max_width=40):
        for col in range(1, col_count + 1):
            letter = get_column_letter(col)
            max_len = 0
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 4, max_width)

    # ========== Sheet 1: Summary ==========
    ws_sum = wb.active
    ws_sum.title = "Summary"

    total = len(rows)
    approved = sum(1 for r in rows if r["approved"])
    mismatches = sum(1 for r in rows if r["was_mismatch"])

    ws_sum.append(["WAF Classification History \u2014 Summary Report"])
    ws_sum.merge_cells("A1:D1")
    ws_sum.cell(1, 1).font = Font(bold=True, size=16, color="1F4E79")

    ws_sum.append([])
    ws_sum.append(["Generated", datetime.now().strftime("%B %d, %Y %I:%M %p")])
    ws_sum.append(["Total Classifications", total])
    ws_sum.append(["Approved to Ground Truth", approved])
    ws_sum.append(["Mismatches Detected", mismatches])
    ws_sum.append(["Approval Rate", f"{round(approved/total*100,1)}%" if total else "0%"])
    ws_sum.append([])

    # Category summary
    ws_sum.append(["WAF Category", "Count", "% of Total"])
    row_num = ws_sum.max_row
    for col in range(1, 4):
        c = ws_sum.cell(row=row_num, column=col)
        c.font = header_font; c.fill = header_fill; c.alignment = header_align

    cat_counts = defaultdict(int)
    for r in rows:
        cat_counts[r["waf_category"] or "Unknown"] += 1

    for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
        ws_sum.append([cat, cnt, f"{round(cnt/total*100,1)}%" if total else "0%"])

    auto_width(ws_sum, 4)

    # ========== Sheet 2: Monthly Rollups ==========
    ws_month = wb.create_sheet("Monthly Rollups")

    monthly = defaultdict(list)
    for r in rows:
        from datetime import datetime as dt
        ts = dt.fromisoformat(r["timestamp"])
        key = ts.strftime("%Y-%m")
        monthly[key].append(r)

    headers_m = ["Month", "Total", "Approved", "Mismatches", "Approval Rate",
                 "Mismatch Rate", "Run", "Change", "Top Category"]
    ws_month.append(headers_m)
    style_header(ws_month, len(headers_m))

    for month_key in sorted(monthly.keys()):
        mrows = monthly[month_key]
        mt = len(mrows)
        ma = sum(1 for r in mrows if r["approved"])
        mm = sum(1 for r in mrows if r["was_mismatch"])
        rc = defaultdict(int)
        mc = defaultdict(int)
        for r in mrows:
            if r["run_change"]: rc[r["run_change"]] += 1
            mc[r["waf_category"] or "Unknown"] += 1
        top_cat = max(mc, key=mc.get) if mc else "\u2014"

        ws_month.append([
            dt.strptime(month_key, "%Y-%m").strftime("%B %Y"),
            mt, ma, mm,
            f"{round(ma/mt*100,1)}%" if mt else "0%",
            f"{round(mm/mt*100,1)}%" if mt else "0%",
            rc.get("Run", 0), rc.get("Change", 0),
            top_cat
        ])

    auto_width(ws_month, len(headers_m))

    # ========== Sheet 3: Raw Data ==========
    ws_raw = wb.create_sheet("Raw Data")

    headers_r = ["Timestamp", "Story Title", "Description", "WAF Category",
                 "Team of Teams", "WAF Color", "Run/Change", "Confidence",
                 "Mismatch", "Original Tag", "Approved", "Team"]
    ws_raw.append(headers_r)
    style_header(ws_raw, len(headers_r))

    for i, r in enumerate(rows, 2):
        ws_raw.append([
            r["timestamp"], r["story_title"], r["story_description"],
            r["waf_category"], r["team_of_teams"], r["waf_color"],
            r["run_change"], r["confidence"],
            "Yes" if r["was_mismatch"] else "No",
            r["original_tag"],
            "Yes" if r["approved"] else "No",
            r["team"]
        ])
        # Conditional formatting
        if r["was_mismatch"]:
            for col in range(1, len(headers_r) + 1):
                ws_raw.cell(row=i, column=col).fill = red_fill
        elif r["approved"]:
            for col in range(1, len(headers_r) + 1):
                ws_raw.cell(row=i, column=col).fill = green_fill

    auto_width(ws_raw, len(headers_r))
    ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(headers_r))}{ws_raw.max_row}"

    # -- Write to buffer --
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=waf-history-report.xlsx"}
    )


@analytics_bp.route("/api/search")
def search_classifications():
    """Full-text search across all classifications using FTS5."""
    q = request.args.get("q", "").strip()
    upload_id = request.args.get("upload_id", "")
    limit = min(int(request.args.get("limit", 25)), 100)

    if len(q) < 2:
        return jsonify({"results": [], "total": 0, "query": q})

    db = get_db()

    # Build FTS5 match term — prefix match on last token, exact on others
    safe_q = q.replace('"', '').replace("'", "")
    fts_term = " ".join(
        (tok + "*" if i == len(safe_q.split()) - 1 else tok)
        for i, tok in enumerate(safe_q.split())
    )

    uid_clause = "AND c.upload_id = ?" if upload_id else ""
    params = [fts_term]
    if upload_id:
        params.append(upload_id)
    params.append(limit)

    try:
        rows = db.execute(f"""
            SELECT c.id, c.story_title, c.waf_category, c.waf_color,
                   c.confidence, c.was_mismatch, c.team, c.epic,
                   c.parent_feature, c.upload_id, c.timestamp,
                   c.story_id, c.feature_id, c.epic_id,
                   u.filename, u.uploaded_at,
                   rank
            FROM classifications c
            LEFT JOIN upload_history u ON c.upload_id = u.id
            JOIN (
                SELECT rowid, rank
                FROM classifications_fts
                WHERE classifications_fts MATCH ?
            ) fts ON fts.rowid = c.id
            WHERE 1=1 {uid_clause}
            ORDER BY fts.rank
            LIMIT ?
        """, params).fetchall()
    except Exception as e:
        return jsonify({"error": str(e), "results": [], "total": 0, "query": q}), 500

    results = []
    for r in rows:
        # Highlight matched term in title
        title = r["story_title"] or ""
        results.append({
            "id": r["id"],
            "title": title,
            "team": r["team"] or "",
            "epic": r["epic"] or "",
            "feature": r["parent_feature"] or "",
            "category": r["waf_category"] or "",
            "color": r["waf_color"] or "",
            "confidence": r["confidence"] or "",
            "is_mismatch": bool(r["was_mismatch"]),
            "upload_id": r["upload_id"],
            "filename": r["filename"] or "Unknown source",
            "date": (r["uploaded_at"] or r["timestamp"] or "")[:10],
            "story_id": r["story_id"] or "",
            "feature_id": r["feature_id"] or "",
            "epic_id": r["epic_id"] or "",
        })

    return jsonify({"results": results, "total": len(results), "query": q})


@analytics_bp.route("/api/narrative", methods=["POST"])
def generate_narrative():
    """Generate an AI narrative summary from WAF classification stats."""
    try:
        from waf_core import get_client, AI_MODEL
        from config import AI_BACKEND
    except ImportError as e:
        return jsonify({"error": f"AI backend unavailable: {e}"}), 503

    data = request.get_json(force=True) or {}
    total      = data.get("total", 0)
    mismatches = data.get("mismatches", 0)
    cats       = data.get("categories", [])
    sub_waf    = data.get("submitted_waf", [])
    run_change = data.get("run_change", [])
    confidence = data.get("confidence", [])

    if total == 0:
        return jsonify({"error": "No data to summarise"}), 400

    mismatch_pct = round(mismatches / total * 100, 1) if total else 0

    # Build a compact stats block for the prompt
    cat_lines = "\n".join(
        f"  - {c['name']}: {c['count']} ({round(c['count']/total*100,1)}%)"
        for c in cats[:8]
    )
    sub_lines = "\n".join(
        f"  - {c['name']}: {c['count']} ({round(c['count']/total*100,1)}%)"
        for c in sub_waf[:6]
    ) if sub_waf else "  (no user tags submitted)"
    rc_lines  = ", ".join(f"{r['name']}: {r['count']} ({round(r['count']/total*100,1)}%)" for r in run_change)
    conf_lines = ", ".join(f"{c['name']}: {round(c['count']/total*100,1)}%" for c in confidence)

    prompt = f"""You are a WAF (Work Allocation Framework) portfolio analyst. Write a concise 3-4 sentence executive narrative summarising the following classification results. Be specific with numbers. Highlight the most important pattern and flag any concerns (high mismatch rate, skewed posture, missing tags, etc.). Write in plain English for a product or engineering leader — no bullet points, no headers, just flowing prose.

STATS:
Total stories: {total}
Mismatch rate: {mismatch_pct}% ({mismatches} stories)
Run / Change: {rc_lines or 'not available'}
AI Confidence: {conf_lines or 'not available'}

AI Suggested WAF Distribution:
{cat_lines}

User Submitted WAF Distribution:
{sub_lines}

Write the narrative now:"""

    try:
        client = get_client()
        if AI_BACKEND == "anthropic":
            response = client.messages.create(
                model=AI_MODEL,
                max_tokens=300,
                messages=[{"role": "user", "content": prompt}]
            )
            narrative = response.content[0].text.strip()
        else:
            # Bedrock
            body = json.dumps({
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 300,
                "messages": [{"role": "user", "content": prompt}]
            })
            resp = client.invoke_model(modelId=AI_MODEL, body=body)
            narrative = json.loads(resp["body"].read())["content"][0]["text"].strip()

        return jsonify({"narrative": narrative})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
